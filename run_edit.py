#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VE3 Tool - MASTER: Edit Mode (Compose MP4)
Quét VISUAL folder và ghép video từ ảnh + voice + SRT.

Usage:
    python run_edit.py                     (quét và xử lý tự động)
    python run_edit.py AR47-0028           (chạy 1 project cụ thể)
    python run_edit.py --parallel 3        (chạy 3 project song song)
    python run_edit.py --scan-only         (chỉ quét, không xử lý)
"""

import sys
import os
import time
import shutil
import json
import re
import unicodedata
import subprocess
import argparse
import random
import tempfile
import gc
import threading
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Set
import queue as _queue_module
import concurrent.futures as _cf
from concurrent.futures import ThreadPoolExecutor, as_completed
from enum import Enum

try:
    from fontTools.ttLib import TTFont
except Exception:
    TTFont = None

# For hardware detection
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False

# Import Ken Burns CV2 module
try:
    from modules.ken_burns_cv2 import KenBurnsCv2, QUALITY_PRESETS
    KEN_BURNS_CV2_AVAILABLE = True
except ImportError:
    KEN_BURNS_CV2_AVAILABLE = False
    print("[WARN] ken_burns_cv2 module not found, using FFmpeg fallback")

# ============================================================================
# CONFIG
# ============================================================================

TOOL_DIR = Path(__file__).parent
VISUAL_DIR = Path(r"D:\AUTO\VISUAL")
DONE_DIR = Path(r"D:\AUTO\done")
THUMB_DIR = Path(r"D:\AUTO\thumbnails")
VOICE_DIR = Path(r"D:\AUTO\voice")  # Voice source folder
PROJECTS_DIR = Path(r"D:\VE3_SUITE\PROJECTS")  # SRT projects folder
CONFIG_FILE = TOOL_DIR / "config" / "config.json"
PROGRESS_FILE = TOOL_DIR / "progress.json"

# Inject ffmpeg path into process environment (supports D:\upload\ffmpeg\bin)
_FFMPEG_CANDIDATES = [
    Path(r"D:\upload\ffmpeg\bin"),
    Path(r"E:\AUTOMATION\ffmpeg\bin"),
    Path(r"C:\ffmpeg\bin"),
    Path(r"D:\ffmpeg\bin"),
]
for _ffdir in _FFMPEG_CANDIDATES:
    if _ffdir.exists() and (_ffdir / "ffmpeg.exe").exists():
        import os as _os
        _os.environ["PATH"] = str(_ffdir) + _os.pathsep + _os.environ.get("PATH", "")
        print(f"[CONFIG] ffmpeg path: {_ffdir}")
        break

SCAN_INTERVAL = 30  # Scan every 30 seconds for new projects
DEFAULT_PARALLEL = 4  # Will be auto-adjusted based on hardware

# Track currently processing codes to prevent duplicate processing
_processing_codes = set()
_processing_lock = threading.Lock()

# ============================================================================
# HARDWARE DETECTION & RESOURCE OPTIMIZATION
# ============================================================================

def detect_system_resources():
    """Detect system hardware resources for optimization."""
    resources = {
        "cpu_cores": os.cpu_count() or 4,
        "cpu_physical": os.cpu_count() or 4,
        "ram_gb": 8,  # Default
        "gpu_available": False,
        "gpu_encoder": "libx264",
        "cpu_percent": 0,
    }

    # Get physical CPU cores (not hyperthreading)
    if PSUTIL_AVAILABLE:
        try:
            resources["cpu_physical"] = psutil.cpu_count(logical=False) or resources["cpu_cores"]
            resources["ram_gb"] = psutil.virtual_memory().total / (1024**3)
            resources["cpu_percent"] = psutil.cpu_percent(interval=0.1)
        except:
            pass

    # Detect GPU encoder
    try:
        gpu_check = subprocess.run(
            ["ffmpeg", "-hide_banner", "-encoders"],
            capture_output=True, text=True, timeout=5,
            creationflags=SUBPROCESS_FLAGS if sys.platform == "win32" else 0
        )
        if "h264_nvenc" in gpu_check.stdout:
            resources["gpu_available"] = True
            resources["gpu_encoder"] = "h264_nvenc"
    except:
        pass

    return resources


def get_optimal_workers(resources: dict = None, task_type: str = "clip") -> int:
    """
    Calculate optimal number of parallel workers based on system resources.

    Args:
        resources: System resources dict from detect_system_resources()
        task_type: "clip" for clip creation, "encode" for encoding, "parallel" for parallel projects

    Returns:
        Optimal number of workers
    """
    if resources is None:
        resources = detect_system_resources()

    cpu_physical = resources.get("cpu_physical", 4)
    ram_gb = resources.get("ram_gb", 8)
    gpu_available = resources.get("gpu_available", False)

    if task_type == "parallel":
        # For parallel video projects:
        # Goal: Maximum throughput (100% CPU is OK)

        # RAM-based: Each project needs ~4GB
        ram_parallel = max(1, int((ram_gb - 8) / 4))

        # CPU-based: Use half of physical cores per project
        cpu_parallel = max(1, cpu_physical // 2)

        # GPU bonus: NVENC allows more parallel work
        if gpu_available:
            gpu_bonus = 2  # Full bonus for speed
        else:
            gpu_bonus = 0

        optimal = min(ram_parallel, cpu_parallel) + gpu_bonus

        # Allow override via environment variable (e.g. VE3_PARALLEL=3 for powerful machines)
        env_parallel = os.environ.get("VE3_PARALLEL")
        if env_parallel:
            return int(env_parallel)
        return 2

    elif task_type == "clip":
        # For OpenCV Ken Burns (CPU intensive):
        # Target ~90% CPU usage

        # RAM-based limit: Each 4K clip needs ~2GB
        ram_workers = max(1, int((ram_gb - 8) / 2))

        # CPU-based limit: Use all physical cores for max throughput
        cpu_workers = cpu_physical

        # Take minimum of constraints
        optimal = min(ram_workers, cpu_workers)

        # Cap workers per video (override via VE3_CLIP_WORKERS for powerful machines)
        env_cap = os.environ.get("VE3_CLIP_WORKERS")
        cap = int(env_cap) if env_cap else 8
        return max(1, min(optimal, cap))

    elif task_type == "encode":
        # For FFmpeg encoding:
        # GPU can handle multiple streams
        if gpu_available:
            return min(4, cpu_physical)  # GPU can handle more parallelism
        else:
            return max(1, cpu_physical // 2)  # CPU encoding is heavy

    return 4  # Default


# Auto-detect resources at startup
_system_resources = None

def get_system_resources():
    """Get cached system resources (detect once)."""
    global _system_resources
    if _system_resources is None:
        _system_resources = detect_system_resources()
    return _system_resources


# Dynamic CLIP_WORKERS based on system resources
CLIP_WORKERS = get_optimal_workers(task_type="clip")  # Will be set based on hardware

# Google Sheet config
SOURCE_SHEET_NAME = "NGUON"
SOURCE_COL_CODE = 7
SOURCE_COL_STATUS = 13
STATUS_VALUE = "EDIT XONG"

MAX_RETRIES = 7  # Increased for Google Sheets reliability
RETRY_BASE_DELAY = 3  # Start with 3s delay

# Hide console window for subprocess on Windows
if sys.platform == "win32":
    SUBPROCESS_FLAGS = subprocess.CREATE_NO_WINDOW
else:
    SUBPROCESS_FLAGS = 0


# Progress tracking for GUI - supports multiple videos in parallel
_multi_progress = {
    "videos": {},      # Dict of code -> video progress
    "hardware": None,  # Hardware info (set at startup)
    "updated": None,
}
_progress_lock = threading.Lock()


def update_progress(code: str = None, step: str = None, percent: int = None,
                   clip_current: int = None, clip_total: int = None, status: str = None,
                   remove: bool = False):
    """Update progress for a specific video and write to file for GUI to read.

    Args:
        code: Video code (required for multi-video tracking)
        step: Current processing step
        percent: Progress percentage (0-100)
        clip_current: Current clip number
        clip_total: Total clips
        status: Status string
        remove: If True, remove this video from tracking (when done)
    """
    from datetime import datetime

    now = datetime.now()

    with _progress_lock:
        # Handle remove request
        if remove and code and code in _multi_progress["videos"]:
            del _multi_progress["videos"][code]
            _write_progress()
            return

        # If no code provided and videos exist, do nothing (avoid overwrite)
        if not code:
            # Legacy support: if no videos, show idle state
            if not _multi_progress["videos"]:
                _multi_progress["updated"] = time.strftime("%H:%M:%S")
                _write_progress()
            return

        # Get or create video progress entry
        if code not in _multi_progress["videos"]:
            _multi_progress["videos"][code] = {
                "code": code,
                "step": "",
                "percent": 0,
                "clip_current": 0,
                "clip_total": 0,
                "status": "starting",
                "started_at": now.isoformat(),
                "step_started_at": now.isoformat(),
                "elapsed_seconds": 0,
                "eta_seconds": None,
            }

        video_progress = _multi_progress["videos"][code]

        # Update step (track timing when step changes)
        if step is not None and step != video_progress.get("step"):
            video_progress["step"] = step
            video_progress["step_started_at"] = now.isoformat()

        # Update other fields
        if percent is not None:
            video_progress["percent"] = percent
        if clip_current is not None:
            video_progress["clip_current"] = clip_current
        if clip_total is not None:
            video_progress["clip_total"] = clip_total
        if status is not None:
            video_progress["status"] = status

        # Calculate elapsed time and ETA
        if video_progress.get("started_at"):
            try:
                started = datetime.fromisoformat(video_progress["started_at"])
                elapsed = (now - started).total_seconds()
                video_progress["elapsed_seconds"] = int(elapsed)

                # Calculate ETA based on progress
                clip_cur = video_progress.get("clip_current", 0)
                clip_tot = video_progress.get("clip_total", 0)
                step_name = video_progress.get("step", "").lower()

                # ETA for clip creation phase
                if "clip" in step_name and clip_tot > 0:
                    step_started = video_progress.get("step_started_at")
                    if step_started and clip_cur > 0:
                        step_start_time = datetime.fromisoformat(step_started)
                        step_elapsed = (now - step_start_time).total_seconds()
                        per_clip = step_elapsed / clip_cur
                        remaining_clips = clip_tot - clip_cur
                        clips_eta = per_clip * remaining_clips
                        video_progress["eta_seconds"] = int(clips_eta * 1.2)
                    elif clip_tot > 0:
                        video_progress["eta_seconds"] = int(clip_tot * 3)

                # General ETA based on percent
                current_percent = video_progress.get("percent", 0)
                if current_percent > 5 and "clip" not in step_name:
                    total_estimated = elapsed / (current_percent / 100)
                    remaining = total_estimated - elapsed
                    video_progress["eta_seconds"] = max(0, int(remaining))
            except:
                pass

        _multi_progress["updated"] = time.strftime("%H:%M:%S")
        _write_progress()


def _write_progress():
    """Write progress to file (must be called with lock held)."""
    try:
        # Build output structure
        output = {
            "videos": _multi_progress["videos"],
            "hardware": _multi_progress.get("hardware"),
            "updated": _multi_progress.get("updated"),
            "active_count": len(_multi_progress["videos"]),
        }
        with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
            json.dump(output, f, indent=2)
    except:
        pass


def set_hardware_info(hardware: dict):
    """Set hardware info in progress."""
    with _progress_lock:
        _multi_progress["hardware"] = hardware
        _write_progress()


def log(msg: str, level: str = "INFO"):
    timestamp = time.strftime("%H:%M:%S")
    line = f"[{timestamp}] [{level}] {msg}"
    try:
        print(line)
    except UnicodeEncodeError:
        import sys as _sys
        encoded = (line + "\n").encode('utf-8', errors='replace')
        _sys.stdout.buffer.write(encoded)
        _sys.stdout.buffer.flush()


VOICE_TARGET_LUFS = -14.0
VOICE_TRUE_PEAK = -1.5
VOICE_TARGET_LRA = 7.0


def _extract_loudnorm_json(stderr: str) -> Optional[dict]:
    match = re.search(r"\{[\s\S]*?\}", stderr or "")
    if not match:
        return None
    try:
        return json.loads(match.group())
    except Exception:
        return None


def measure_audio_loudness(audio_path: Path, dual_mono: bool = True, pre_filter: str = "") -> Optional[dict]:
    """Measure integrated loudness/true peak using FFmpeg loudnorm JSON output."""
    loudnorm = (
        f"loudnorm=I={VOICE_TARGET_LUFS}:TP={VOICE_TRUE_PEAK}:LRA={VOICE_TARGET_LRA}"
        f":dual_mono={'true' if dual_mono else 'false'}:print_format=json"
    )
    af = f"{pre_filter},{loudnorm}" if pre_filter else loudnorm
    cmd = ["ffmpeg", "-hide_banner", "-nostats", "-i", str(audio_path), "-af", af, "-f", "null", "-"]
    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=900,
            creationflags=SUBPROCESS_FLAGS
        )
    except Exception:
        return None
    stats = _extract_loudnorm_json(result.stderr)
    if stats is not None:
        stats["returncode"] = result.returncode
    return stats


def _voice_pre_filter() -> str:
    # Mild speech cleanup + local leveling before the final LUFS pass.
    return (
        "highpass=f=70,"
        "lowpass=f=14500,"
        "dynaudnorm=f=750:g=15:p=0.90:m=8:s=6:t=0.02"
    )


def normalize_voice(voice_path: Path, temp_dir: Path, label: str = "voice") -> Path:
    """
    Speech-first mastering for ElevenLabs voice.

    Order matters: local leveling/filtering first, then 2-pass loudnorm last.
    This keeps merged chunks even while preserving final YouTube-ready LUFS/true peak.
    """
    out_path = temp_dir / f"{label}_norm_{voice_path.stem}.mp3"
    report_path = temp_dir / f"{label}_audio_report_{voice_path.stem}.json"
    pre_filter = _voice_pre_filter()

    before_stats = measure_audio_loudness(voice_path, dual_mono=False)
    measured = measure_audio_loudness(voice_path, dual_mono=False, pre_filter=pre_filter)
    if not measured:
        log(f"Audio measure failed for {voice_path.name}; trying fallback leveling", "WARN")
    else:
        try:
            af = (
                f"{pre_filter},"
                f"loudnorm=I={VOICE_TARGET_LUFS}:TP={VOICE_TRUE_PEAK}:LRA={VOICE_TARGET_LRA}"
                f":dual_mono=false"
                f":measured_I={measured['input_i']}"
                f":measured_TP={measured['input_tp']}"
                f":measured_LRA={measured['input_lra']}"
                f":measured_thresh={measured['input_thresh']}"
                f":offset={measured['target_offset']}"
                f":linear=false"
            )
            cmd = [
                "ffmpeg", "-y", "-hide_banner", "-nostats", "-i", str(voice_path),
                "-af", af, "-ar", "44100", "-ac", "2", "-b:a", "256k",
                str(out_path)
            ]
            result = subprocess.run(
                cmd, capture_output=True, text=True, timeout=900,
                creationflags=SUBPROCESS_FLAGS
            )
            if result.returncode == 0 and out_path.exists() and out_path.stat().st_size > 1000:
                after_stats = measure_audio_loudness(out_path, dual_mono=False)
                try:
                    report = {
                        "source": str(voice_path),
                        "output": str(out_path),
                        "target_lufs": VOICE_TARGET_LUFS,
                        "target_true_peak": VOICE_TRUE_PEAK,
                        "target_lra": VOICE_TARGET_LRA,
                        "before": before_stats,
                        "measured_after_prefilter": measured,
                        "after": after_stats,
                    }
                    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
                except Exception:
                    pass
                if after_stats:
                    log(
                        f"Audio mastered {voice_path.name}: "
                        f"I={after_stats.get('input_i')} LUFS, "
                        f"TP={after_stats.get('input_tp')} dBTP, "
                        f"LRA={after_stats.get('input_lra')}"
                    )
                return out_path
            log(f"Audio normalize failed: {result.stderr[-220:]}", "WARN")
        except Exception as exc:
            log(f"Audio normalize error: {exc}", "WARN")

    # Fallback keeps speech intelligible even if 2-pass loudnorm cannot parse stats.
    cmd_fb = [
        "ffmpeg", "-y", "-hide_banner", "-nostats", "-i", str(voice_path),
        "-af", f"{pre_filter},alimiter=limit=0.841:level=false",
        "-ar", "44100", "-ac", "2", "-b:a", "256k",
        str(out_path)
    ]
    result_fb = subprocess.run(
        cmd_fb, capture_output=True, text=True, timeout=900,
        creationflags=SUBPROCESS_FLAGS
    )
    if result_fb.returncode == 0 and out_path.exists() and out_path.stat().st_size > 1000:
        return out_path

    return voice_path


def save_timing_log(entry: dict):
    """Lưu nhật ký thời gian từng bước vào timing_log.json để phân tích tối ưu."""
    import json
    log_path = TOOL_DIR / "timing_log.json"
    entries = []
    if log_path.exists():
        try:
            entries = json.loads(log_path.read_text(encoding="utf-8"))
        except Exception:
            entries = []
    entries.append(entry)
    log_path.write_text(json.dumps(entries, indent=2, ensure_ascii=False), encoding="utf-8")


# ============================================================================
# SUBTITLE TEMPLATE SYSTEM
# ============================================================================

SUBTITLE_TEMPLATES_FILE = TOOL_DIR / "subtitle_templates.json"

# Default template (used when no channel-specific template exists)
DEFAULT_SUBTITLE_TEMPLATE = {
    "font": "Auto",
    "size": 28,
    "color": "&H00FFFFFF",      # White (ABGR format)
    "outline": "&H00000000",    # Black outline
    "outline_size": 2,
    "margin_v": 25,
    "alignment": 2,             # 2 = bottom center
    # Video settings (per channel)
    "output_resolution": "4k",
    "compose_mode": "quality",
    "ken_burns_intensity": "subtle",
    "video_transition": "random",
    # Subtitle timing sync target:
    # - "voice" (default): keep subtitle aligned with spoken audio
    # - "video": compensate subtitle timestamps for xfade timeline compression
    "subtitle_sync": "voice",
    # NV overlay settings
    "nv_overlay_enabled": False,
    "nv_overlay_position": "left",
    "nv_overlay_v_position": "middle",
    "nv_overlay_scale": 0.50,
    "nv_crop_ratio": 0.5,  # Crop right portion (0.5 = right half, 1.0 = full image)
    # Remove Veo watermark on source videos via minimal corner crop (before upscale)
    "veo_crop_enabled": True,
    "veo_crop_right_ratio": 0.05,
    "veo_crop_bottom_ratio": 0.07,
    "veo_crop_min_right_px": 56,
    "veo_crop_min_bottom_px": 48,
    "veo_crop_keep_4k_aspect": True,
}

# Available fonts in fonts/ folder
AVAILABLE_FONTS = [
    "Auto",
    "Bebas Neue",
    "Oswald",
    "Roboto Condensed",
    "Noto Sans Condensed",
    "Noto Sans",
    "Noto Serif",
    "Inter",
    "Inter 18pt",
    "Anton",
    "League Spartan ExtraBold",
    "Montserrat Thin",
    "Nunito ExtraLight",
    "Roboto Condensed Black",
    "UTM Avo",
    "Zuume SemiBold"
]

# Font names must match the internal family name that libass sees, not only the
# .ttf file name shown in Explorer. Keep common GUI/file-name aliases here.
SUBTITLE_FONT_ALIASES = {
    "auto": "Auto",
    "smart": "Auto",
    "tu dong": "Auto",
    "tự động": "Auto",
    "bebasneue": "Bebas Neue",
    "bebasneue regular": "Bebas Neue",
    "bebas neue": "Bebas Neue",
    "bebas neue regular": "Bebas Neue",
    "notoserif": "Noto Serif",
    "notoserif regular": "Noto Serif",
    "noto serif": "Noto Serif",
    "noto serif regular": "Noto Serif",
    "notosans": "Noto Sans",
    "notosans regular": "Noto Sans",
    "noto sans": "Noto Sans",
    "noto sans regular": "Noto Sans",
    "notosans bold": "Noto Sans",
    "noto sans bold": "Noto Sans",
    "notosans condensedbold": "Noto Sans Condensed",
    "notosans condensed bold": "Noto Sans Condensed",
    "noto sans condensed": "Noto Sans Condensed",
    "noto sans condensed bold": "Noto Sans Condensed",
    "inter": "Inter",
    "inter bold": "Inter 18pt",
    "inter 18pt": "Inter 18pt",
    "anton": "Anton",
    "league spartan": "League Spartan ExtraBold",
    "league spartan extrabold": "League Spartan ExtraBold",
    "montserrat": "Montserrat Thin",
    "nunito": "Nunito ExtraLight",
    "oswald": "Oswald",
    "oswald wght": "Oswald",
    "roboto condensed": "Roboto Condensed",
    "roboto condensed regular": "Roboto Condensed",
    "robotocondensed": "Roboto Condensed",
    "robotocondensed regular": "Roboto Condensed",
    "robotocondensed wght": "Roboto Condensed",
    "roboto condensed black": "Roboto Condensed Black",
    "utm avo": "UTM Avo",
    "utm avo bold": "UTM Avo",
    "utm-avobold": "UTM Avo",
    "ut mavobold": "UTM Avo",
    "zuume semibold": "Zuume SemiBold",
}

SUBTITLE_SMART_FONT_PRIORITY = [
    "Bebas Neue",
    "Oswald",
    "Roboto Condensed",
    "Noto Sans Condensed",
    "Anton",
    "Noto Sans",
    "Noto Serif",
    "Inter",
    "Inter 18pt",
    "Nunito ExtraLight",
    "Montserrat Thin",
    "Roboto Condensed Black",
    "UTM Avo",
    "Arial",
]

SUBTITLE_LANGUAGE_FONT_PROFILES = {
    # Keep the strong, clear YouTube-subtitle look first; fallback fonts are
    # chosen for complete Latin Extended coverage.
    "vi": ["Bebas Neue", "Oswald", "Roboto Condensed", "Noto Sans Condensed", "UTM Avo", "Noto Sans", "Inter 18pt"],
    "es": ["Bebas Neue", "Oswald", "Roboto Condensed", "Noto Sans Condensed", "Anton", "Noto Sans", "Inter 18pt"],
    "en": ["Bebas Neue", "Anton", "Oswald", "Roboto Condensed", "Noto Sans Condensed", "Inter 18pt", "Noto Sans"],
    "fr": ["Bebas Neue", "Oswald", "Roboto Condensed", "Noto Sans Condensed", "Noto Sans", "Inter 18pt", "Noto Serif"],
    "de": ["Bebas Neue", "Oswald", "Roboto Condensed", "Noto Sans Condensed", "Noto Sans", "Inter 18pt", "Noto Serif"],
    "pt": ["Bebas Neue", "Oswald", "Roboto Condensed", "Noto Sans Condensed", "Noto Sans", "Inter 18pt", "Noto Serif"],
    "latin": SUBTITLE_SMART_FONT_PRIORITY,
}

SUBTITLE_RENDER_PROFILES = {
    "UTM Avo": {
        "size_scale": 0.60,
        "max_chars": 72,
        "outline_scale": 0.55,
        "margin_v_delta": 8,
    },
    "Oswald": {
        "size_scale": 0.72,
        "max_chars": 72,
        "outline_scale": 0.65,
        "margin_v_delta": 6,
    },
    "Roboto Condensed": {
        "size_scale": 0.90,
        "max_chars": 38,
        "outline_scale": 1.00,
        "margin_v_delta": 6,
    },
    "Noto Sans Condensed": {
        "size_scale": 0.90,
        "max_chars": 38,
        "outline_scale": 1.00,
        "margin_v_delta": 6,
    },
}

_FONT_SCAN_CACHE = None

# Alignment options: 1=left, 2=center, 3=right (bottom row)
# 4=left, 5=center, 6=right (middle row)
# 7=left, 8=center, 9=right (top row)
ALIGNMENT_OPTIONS = {
    "bottom_left": 1,
    "bottom_center": 2,
    "bottom_right": 3,
    "middle_left": 4,
    "middle_center": 5,
    "middle_right": 6,
    "top_left": 7,
    "top_center": 8,
    "top_right": 9
}


def load_subtitle_templates() -> Dict:
    """Load subtitle templates from JSON file."""
    if SUBTITLE_TEMPLATES_FILE.exists():
        try:
            with open(SUBTITLE_TEMPLATES_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return {}


def save_subtitle_templates(templates: Dict):
    """Save subtitle templates to JSON file."""
    try:
        with open(SUBTITLE_TEMPLATES_FILE, "w", encoding="utf-8") as f:
            json.dump(templates, f, indent=2, ensure_ascii=False)
    except Exception as e:
        log(f"Error saving templates: {e}", "ERROR")


def get_subtitle_template(code: str) -> Dict:
    """Get subtitle template for a channel based on code prefix (e.g., KA1, KA2)."""
    templates = load_subtitle_templates()

    # Extract channel prefix (e.g., "KA1" from "KA1-0023")
    channel = code.split("-")[0] if "-" in code else code

    # Look for exact channel match
    if channel in templates:
        template = DEFAULT_SUBTITLE_TEMPLATE.copy()
        template.update(templates[channel])
        return template

    # Look for base channel (e.g., "KA" from "KA1")
    base_channel = ''.join(c for c in channel if not c.isdigit())
    if base_channel in templates:
        template = DEFAULT_SUBTITLE_TEMPLATE.copy()
        template.update(templates[base_channel])
        return template

    return DEFAULT_SUBTITLE_TEMPLATE.copy()


def set_subtitle_template(channel: str, template: Dict):
    """Set subtitle template for a channel."""
    templates = load_subtitle_templates()
    templates[channel] = template
    save_subtitle_templates(templates)
    log(f"Saved template for channel: {channel}")


def get_all_templates() -> Dict:
    """Get all saved templates."""
    return load_subtitle_templates()


def resolve_subtitle_font(font_name: str) -> str:
    """Return a libass-friendly font family name."""
    raw = str(font_name or "").strip()
    key = raw.lower().replace("_", " ").replace("-", " ")
    key = re.sub(r"\s+", " ", key).strip()
    compact_key = key.replace(" ", "")
    return SUBTITLE_FONT_ALIASES.get(key) or SUBTITLE_FONT_ALIASES.get(compact_key) or raw or "Noto Serif"


def subtitle_required_codepoints(text: str) -> Set[int]:
    """Codepoints that a subtitle font must render."""
    required = set()
    for ch in text or "":
        if ch.isspace() or unicodedata.category(ch).startswith("C"):
            continue
        required.add(ord(ch))
    return required


def scan_subtitle_font_files() -> Dict[str, Dict]:
    """Scan fonts/ and cache internal family names plus glyph coverage."""
    global _FONT_SCAN_CACHE
    if _FONT_SCAN_CACHE is not None:
        return _FONT_SCAN_CACHE

    fonts = {}
    if TTFont is None:
        _FONT_SCAN_CACHE = fonts
        return fonts

    fonts_dir = TOOL_DIR / "fonts"
    if not fonts_dir.exists():
        _FONT_SCAN_CACHE = fonts
        return fonts

    for path in fonts_dir.glob("*.ttf"):
        try:
            font = TTFont(path)
            names = set()
            for record in font["name"].names:
                if record.nameID in (1, 4, 6):
                    try:
                        value = record.toUnicode().strip()
                    except Exception:
                        continue
                    if value:
                        names.add(value)

            coverage = set()
            for table in font["cmap"].tables:
                coverage.update(table.cmap.keys())

            aliases = {path.stem}
            for name in names:
                aliases.add(name)
                aliases.add(resolve_subtitle_font(name))
            aliases.add(resolve_subtitle_font(path.stem))

            for alias in aliases:
                resolved = resolve_subtitle_font(alias)
                if not resolved or resolved == "Auto":
                    continue
                fonts.setdefault(resolved, {
                    "path": path,
                    "coverage": coverage,
                    "names": sorted(names),
                })
        except Exception as e:
            log(f"Could not inspect font {path.name}: {e}", "WARN")

    _FONT_SCAN_CACHE = fonts
    return fonts


def font_supports_codepoints(font_name: str, codepoints: Set[int]) -> bool:
    """True when a local font covers all required subtitle glyphs."""
    if not codepoints:
        return True
    fonts = scan_subtitle_font_files()
    info = fonts.get(resolve_subtitle_font(font_name))
    if not info:
        return False
    return codepoints.issubset(info["coverage"])


def detect_subtitle_language(text: str) -> str:
    """Detect the main Latin-script subtitle language for font profiling."""
    sample = unicodedata.normalize("NFC", text or "").lower()
    words = re.findall(r"[a-zà-ỹñçäöüßœæ]+", sample, flags=re.IGNORECASE)
    word_set = set(words)

    scores = {
        "vi": 0,
        "es": 0,
        "fr": 0,
        "de": 0,
        "pt": 0,
        "en": 0,
    }

    vi_chars = set("ăâđêôơưạảãáàắằẳẵặấầẩẫậẹẻẽéèếềểễệịỉĩíìọỏõóòốồổỗộớờởỡợụủũúùứừửữựỵỷỹýỳ")
    es_chars = set("¿¡ñ")
    de_chars = set("ßäöü")
    pt_chars = set("ãõ")
    fr_chars = set("œæêëîïûùç")

    scores["vi"] += min(12, sum(1 for ch in sample if ch in vi_chars))
    scores["es"] += 6 if any(ch in sample for ch in es_chars) else 0
    scores["de"] += min(10, sum(2 for ch in sample if ch in de_chars))
    scores["pt"] += min(10, sum(3 for ch in sample if ch in pt_chars))
    scores["fr"] += min(8, sum(2 for ch in sample if ch in fr_chars))

    keyword_sets = {
        "vi": {"và", "của", "là", "không", "người", "một", "tôi", "bạn", "này", "đó"},
        "es": {"el", "la", "los", "las", "que", "de", "para", "con", "una", "por", "es"},
        "fr": {"le", "la", "les", "des", "que", "pour", "avec", "une", "est", "dans", "pas"},
        "de": {"der", "die", "das", "und", "ist", "nicht", "mit", "für", "ein", "eine", "ich"},
        "pt": {"que", "não", "uma", "para", "com", "por", "está", "você", "mais", "como"},
        "en": {"the", "and", "you", "that", "with", "this", "not", "are", "for", "have"},
    }
    for lang, keywords in keyword_sets.items():
        scores[lang] += len(word_set & keywords)

    lang, score = max(scores.items(), key=lambda item: item[1])
    return lang if score > 0 else "latin"


def choose_subtitle_font(font_name: str, subtitle_text: str) -> str:
    """Choose a subtitle font that can render every glyph in the SRT."""
    requested = resolve_subtitle_font(font_name)
    required = subtitle_required_codepoints(subtitle_text)

    if requested != "Auto" and font_supports_codepoints(requested, required):
        return requested

    detected_language = detect_subtitle_language(subtitle_text)
    profile = SUBTITLE_LANGUAGE_FONT_PROFILES.get(detected_language, SUBTITLE_SMART_FONT_PRIORITY)
    for candidate in profile:
        if font_supports_codepoints(candidate, required):
            return candidate

    for candidate in SUBTITLE_SMART_FONT_PRIORITY:
        if font_supports_codepoints(candidate, required):
            return candidate

    return "Noto Sans" if requested == "Auto" else requested


def get_subtitle_render_profile(font_name: str, detected_language: str) -> Dict:
    """Return sizing/wrapping adjustments for fonts with wider glyph metrics."""
    profile = {
        "size_scale": 1.0,
        "max_chars": 45,
        "outline_scale": 1.0,
        "margin_v_delta": 0,
    }
    font_profile = SUBTITLE_RENDER_PROFILES.get(resolve_subtitle_font(font_name), {})
    profile.update(font_profile)
    if detected_language == "vi":
        profile["max_chars"] = min(profile["max_chars"], 72)
    return profile


def subtitle_style_numbers(template: Dict, font_name: str, detected_language: str) -> Tuple[int, int, int]:
    """Compute ASS style numbers after font/language-specific adjustments."""
    profile = get_subtitle_render_profile(font_name, detected_language)
    base_size = max(10, _to_int(template.get("size", 28), 28))
    base_outline = max(0, _to_int(template.get("outline_size", 2), 2))
    base_margin = max(0, _to_int(template.get("margin_v", 25), 25))
    font_size = max(16, int(round(base_size * profile["size_scale"])))
    outline_size = max(1, int(round(base_outline * profile["outline_scale"]))) if base_outline > 0 else 0
    margin_v = max(0, base_margin + int(profile["margin_v_delta"]))
    return font_size, outline_size, margin_v


def escape_ass_style_value(value) -> str:
    """Keep force_style values from breaking the subtitles filter syntax."""
    return str(value).replace(",", " ").replace("'", "").strip()


def read_text_best_effort(path: Path) -> str:
    """Read subtitle text even when the source file is not UTF-8."""
    data = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "utf-16", "cp1258", "cp1252"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def normalize_srt_utf8(srt_path: Path, output_path: Path) -> Path:
    """Write an UTF-8 SRT copy for ffmpeg/libass."""
    try:
        text = read_text_best_effort(srt_path)
        text = unicodedata.normalize("NFC", text).replace("\r\n", "\n").replace("\r", "\n")
        output_path.write_text(text, encoding="utf-8")
        return output_path
    except Exception as e:
        log(f"Could not normalize SRT encoding: {e}", "WARN")
        return srt_path


def srt_has_cues(srt_path: Path) -> bool:
    """True when an SRT file is non-empty and contains at least one timestamp cue."""
    try:
        if not srt_path.exists() or srt_path.stat().st_size == 0:
            return False
        text = read_text_best_effort(srt_path)
        return bool(re.search(
            r"\d{2}:\d{2}:\d{2}[,\.]\d{3}\s*-->\s*\d{2}:\d{2}:\d{2}[,\.]\d{3}",
            text
        ))
    except Exception:
        return False


def normalize_code(code: str) -> str:
    if not code:
        return ""
    s = str(code)
    s = s.replace("–", "-").replace("—", "-").replace("−", "-")
    s = re.sub(r"\s+", " ", s).strip()
    return s.upper()


def _to_bool(v, default=False) -> bool:
    """Safe bool conversion for template values."""
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in ("1", "true", "yes", "on")
    if v is None:
        return default
    return bool(v)


def _to_float(v, default: float) -> float:
    try:
        return float(v)
    except Exception:
        return float(default)


def _to_int(v, default: int) -> int:
    try:
        return int(v)
    except Exception:
        return int(default)


def compute_veo_crop_geometry(src_w: int, src_h: int, target_w: int, target_h: int,
                              right_ratio: float = 0.03, bottom_ratio: float = 0.02,
                              min_right_px: int = 28, min_bottom_px: int = 12,
                              keep_target_aspect: bool = True) -> Optional[Dict[str, int]]:
    """
    Compute minimal crop from right-bottom corner to remove Veo watermark.

    Returns:
        dict with crop_w/crop_h/crop_x/crop_y/right_crop/bottom_crop
        or None when geometry is invalid.
    """
    try:
        src_w = int(src_w)
        src_h = int(src_h)
        target_w = int(target_w)
        target_h = int(target_h)
    except Exception:
        return None

    if src_w < 16 or src_h < 16:
        return None

    right_ratio = max(0.0, float(right_ratio))
    bottom_ratio = max(0.0, float(bottom_ratio))
    min_right_px = max(2, int(min_right_px))
    min_bottom_px = max(2, int(min_bottom_px))

    req_right = max(min_right_px, int(round(src_w * right_ratio)))
    req_bottom = max(min_bottom_px, int(round(src_h * bottom_ratio)))

    # Keep a safety margin so crop size stays valid
    req_right = min(req_right, max(2, src_w - 8))
    req_bottom = min(req_bottom, max(2, src_h - 8))

    # Minimal crop first: just remove watermark area from right-bottom corner
    crop_w = max(8, (src_w - req_right) - ((src_w - req_right) % 2))
    crop_h = max(8, (src_h - req_bottom) - ((src_h - req_bottom) % 2))

    if keep_target_aspect and target_w > 0 and target_h > 0:
        aspect = target_w / target_h
        base_area = crop_w * crop_h
        if base_area > 0:
            # Candidate A: additional crop on width to match target aspect
            cand_w = int(crop_h * aspect)
            cand_w = max(8, cand_w - (cand_w % 2))
            if cand_w > crop_w:
                cand_w = crop_w
            area_a = cand_w * crop_h

            # Candidate B: additional crop on height to match target aspect
            cand_h = int(crop_w / aspect)
            cand_h = max(8, cand_h - (cand_h % 2))
            if cand_h > crop_h:
                cand_h = crop_h
            area_b = crop_w * cand_h

            # Keep the option with less extra crop (larger kept area)
            if area_a >= area_b:
                adj_w, adj_h, adj_area = cand_w, crop_h, area_a
            else:
                adj_w, adj_h, adj_area = crop_w, cand_h, area_b

            # Always enforce target aspect to avoid padding/letterbox later.
            # Choose the option that keeps more image area.
            crop_w, crop_h = adj_w, adj_h

    crop_w = min(crop_w, src_w - 2)
    crop_h = min(crop_h, src_h - 2)

    if crop_w <= 0 or crop_h <= 0:
        return None

    right_crop = src_w - crop_w
    bottom_crop = src_h - crop_h
    if right_crop < 1 or bottom_crop < 1:
        return None

    return {
        "crop_w": int(crop_w),
        "crop_h": int(crop_h),
        "crop_x": 0,
        "crop_y": 0,
        "right_crop": int(right_crop),
        "bottom_crop": int(bottom_crop),
    }


# ============================================================================
# KEN BURNS EFFECT (inline from modules/ken_burns.py)
# ============================================================================

class KenBurnsEffect(Enum):
    ZOOM_IN = "zoom_in"
    ZOOM_OUT = "zoom_out"
    PAN_LEFT = "pan_left"
    PAN_RIGHT = "pan_right"
    PAN_UP = "pan_up"
    PAN_DOWN = "pan_down"
    ZOOM_IN_LEFT = "zoom_in_left"
    ZOOM_IN_RIGHT = "zoom_in_right"
    ZOOM_OUT_CENTER = "zoom_out_center"


class KenBurnsIntensity(Enum):
    SUBTLE = "subtle"
    NORMAL = "normal"
    STRONG = "strong"


INTENSITY_SETTINGS = {
    KenBurnsIntensity.SUBTLE: (0.05, 0.03),
    KenBurnsIntensity.NORMAL: (0.12, 0.08),
    KenBurnsIntensity.STRONG: (0.20, 0.15),
}


class KenBurnsGenerator:
    def __init__(self, width: int = 1920, height: int = 1080,
                 intensity: str = "normal", fps: int = 25):
        self.width = width
        self.height = height
        self.fps = fps

        if isinstance(intensity, str):
            intensity = intensity.lower()
            self.intensity = {
                "subtle": KenBurnsIntensity.SUBTLE,
                "normal": KenBurnsIntensity.NORMAL,
                "strong": KenBurnsIntensity.STRONG,
            }.get(intensity, KenBurnsIntensity.NORMAL)
        else:
            self.intensity = intensity

        self.zoom_percent, self.pan_percent = INTENSITY_SETTINGS[self.intensity]

    def get_random_effect(self, exclude_last=None):
        effects = list(KenBurnsEffect)
        if exclude_last and exclude_last in effects:
            effects.remove(exclude_last)
        return random.choice(effects)

    def generate_filter(self, effect, duration: float,
                       fade_duration: float = 0.5, simple_mode: bool = False) -> str:
        w, h = self.width, self.height
        total_frames = int(duration * self.fps)

        zoom_start = 1.0
        zoom_end = 1.0 + self.zoom_percent
        pan_x = int(w * self.pan_percent)
        pan_y = int(h * self.pan_percent)

        if simple_mode:
            zoom_expr, x_expr, y_expr = self._get_linear_expressions(
                effect, zoom_start, zoom_end, pan_x, pan_y, total_frames
            )
        else:
            zoom_expr, x_expr, y_expr = self._get_eased_expressions(
                effect, zoom_start, zoom_end, pan_x, pan_y, total_frames
            )

        zoompan = f"zoompan=z='{zoom_expr}':x='{x_expr}':y='{y_expr}':d={total_frames}:s={w}x{h}:fps={self.fps}"
        fade_out_start = max(0, duration - fade_duration)
        fade_filter = f"fade=t=in:st=0:d={fade_duration},fade=t=out:st={fade_out_start}:d={fade_duration}"

        return f"{zoompan},{fade_filter}"

    def _get_linear_expressions(self, effect, zoom_start, zoom_end, pan_x, pan_y, total_frames):
        progress = f"on/{total_frames}"

        if effect == KenBurnsEffect.ZOOM_IN:
            zoom = f"{zoom_start}+{zoom_end - zoom_start}*{progress}"
            x = f"iw/2-(iw/zoom/2)"
            y = f"ih/2-(ih/zoom/2)"
        elif effect == KenBurnsEffect.ZOOM_OUT:
            zoom = f"{zoom_end}-{zoom_end - zoom_start}*{progress}"
            x = f"iw/2-(iw/zoom/2)"
            y = f"ih/2-(ih/zoom/2)"
        elif effect == KenBurnsEffect.PAN_LEFT:
            zoom = str(zoom_start)
            x = f"{pan_x}*(1-{progress})"
            y = "0"
        elif effect == KenBurnsEffect.PAN_RIGHT:
            zoom = str(zoom_start)
            x = f"{pan_x}*{progress}"
            y = "0"
        elif effect == KenBurnsEffect.PAN_UP:
            zoom = str(zoom_start)
            x = "0"
            y = f"{pan_y}*(1-{progress})"
        elif effect == KenBurnsEffect.PAN_DOWN:
            zoom = str(zoom_start)
            x = "0"
            y = f"{pan_y}*{progress}"
        elif effect == KenBurnsEffect.ZOOM_IN_LEFT:
            zoom = f"{zoom_start}+{zoom_end - zoom_start}*{progress}"
            x = f"(iw/4)*(1-{progress})"
            y = f"ih/2-(ih/zoom/2)"
        elif effect == KenBurnsEffect.ZOOM_IN_RIGHT:
            zoom = f"{zoom_start}+{zoom_end - zoom_start}*{progress}"
            x = f"iw/2-(iw/zoom/2)+{pan_x}*{progress}"
            y = f"ih/2-(ih/zoom/2)"
        else:
            zoom = f"{zoom_end}-{zoom_end - zoom_start}*{progress}"
            x = f"iw/2-(iw/zoom/2)"
            y = f"ih/2-(ih/zoom/2)"

        return zoom, x, y

    def _get_eased_expressions(self, effect, zoom_start, zoom_end, pan_x, pan_y, total_frames):
        progress = f"(1-cos(PI*on/{total_frames}))/2"

        if effect == KenBurnsEffect.ZOOM_IN:
            zoom = f"{zoom_start}+{zoom_end - zoom_start}*{progress}"
            x = f"iw/2-(iw/zoom/2)"
            y = f"ih/2-(ih/zoom/2)"
        elif effect == KenBurnsEffect.ZOOM_OUT:
            zoom = f"{zoom_end}-{zoom_end - zoom_start}*{progress}"
            x = f"iw/2-(iw/zoom/2)"
            y = f"ih/2-(ih/zoom/2)"
        elif effect == KenBurnsEffect.PAN_LEFT:
            zoom = str(zoom_start)
            x = f"{pan_x}*(1-{progress})"
            y = "0"
        elif effect == KenBurnsEffect.PAN_RIGHT:
            zoom = str(zoom_start)
            x = f"{pan_x}*{progress}"
            y = "0"
        elif effect == KenBurnsEffect.PAN_UP:
            zoom = str(zoom_start)
            x = "0"
            y = f"{pan_y}*(1-{progress})"
        elif effect == KenBurnsEffect.PAN_DOWN:
            zoom = str(zoom_start)
            x = "0"
            y = f"{pan_y}*{progress}"
        elif effect == KenBurnsEffect.ZOOM_IN_LEFT:
            zoom = f"{zoom_start}+{zoom_end - zoom_start}*{progress}"
            x = f"(iw/4)*(1-{progress})"
            y = f"ih/2-(ih/zoom/2)"
        elif effect == KenBurnsEffect.ZOOM_IN_RIGHT:
            zoom = f"{zoom_start}+{zoom_end - zoom_start}*{progress}"
            x = f"iw/2-(iw/zoom/2)+{pan_x}*{progress}"
            y = f"ih/2-(ih/zoom/2)"
        else:
            zoom = f"{zoom_end}-{zoom_end - zoom_start}*{progress}"
            x = f"iw/2-(iw/zoom/2)"
            y = f"ih/2-(ih/zoom/2)"

        return zoom, x, y


# ============================================================================
# FILL MISSING MEDIA
# ============================================================================

def get_required_scene_ids(excel_path: Path) -> Set[str]:
    """Get all scene IDs required from Excel."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)

        # Find sheet with scene data (must have more than just header row)
        # Priority: sheets with 'scene' in name, then 'srt_coverage' as fallback
        scenes_sheet = None
        fallback_sheet = None
        for sheet_name in wb.sheetnames:
            candidate = wb[sheet_name]
            if candidate.max_row <= 1:
                continue
            if 'scene' in sheet_name.lower() and sheet_name.lower() != 'srt_coverage':
                scenes_sheet = candidate
                break
            elif sheet_name.lower() == 'srt_coverage':
                fallback_sheet = candidate

        if not scenes_sheet and fallback_sheet:
            scenes_sheet = fallback_sheet

        if not scenes_sheet:
            return set()

        headers = [cell.value for cell in scenes_sheet[1]]
        id_col = None
        for i, h in enumerate(headers):
            if h and str(h).lower().strip() in ['scene_id', 'id']:
                id_col = i
                break

        if id_col is None:
            return set()

        scene_ids = set()
        for row in scenes_sheet.iter_rows(min_row=2, values_only=True):
            if row[id_col] is not None:
                try:
                    scene_id = str(int(float(str(row[id_col]).strip())))
                    scene_ids.add(scene_id)
                except ValueError:
                    continue
        return scene_ids
    except Exception as e:
        log(f"Error reading Excel: {e}", "WARN")
        return set()


def get_existing_media(img_dir: Path) -> Dict[str, Path]:
    """Get existing media files mapped by scene ID."""
    media = {}
    if not img_dir.exists():
        return media

    for ext in [".mp4", ".png", ".jpg", ".jpeg", ".webp"]:
        for f in img_dir.glob(f"*{ext}"):
            if f.stem.startswith('nv') or f.stem.startswith('loc'):
                continue
            scene_id = f.stem
            if scene_id not in media:
                media[scene_id] = f
    return media


def fill_missing_media(project_dir: Path, excel_path: Path) -> Tuple[int, int]:
    """Fill missing media by copying random existing media files."""
    # Support both VM structure (img\ subfolder) and direct structure
    img_dir = project_dir / "img"
    if not img_dir.exists():
        img_dir = project_dir  # Fallback to root folder

    # Check if there are any media files
    has_media = any(img_dir.glob("*.mp4")) or any(img_dir.glob("*.png")) or any(img_dir.glob("*.jpg"))
    if not has_media:
        log(f"  [FILL] No media found in project", "WARN")
        return 0, 0

    required_ids = get_required_scene_ids(excel_path)
    if not required_ids:
        log(f"  [FILL] No scenes found in Excel", "WARN")
        return 0, 0

    existing_media = get_existing_media(img_dir)
    existing_ids = set(existing_media.keys())
    missing_ids = required_ids - existing_ids

    if not missing_ids:
        log(f"  [FILL] All {len(required_ids)} scenes have media")
        return 0, 0

    if not existing_media:
        log(f"  [FILL] No existing media to copy from!", "ERROR")
        return 0, len(missing_ids)

    log(f"  [FILL] Missing {len(missing_ids)} scenes, filling from {len(existing_media)} existing...")

    existing_files = list(existing_media.values())
    filled_count = 0

    for missing_id in sorted(missing_ids, key=lambda x: int(x) if x.isdigit() else 0):
        source_file = random.choice(existing_files)
        dest_file = img_dir / f"{missing_id}{source_file.suffix}"

        try:
            shutil.copy2(source_file, dest_file)
            log(f"    Copied {source_file.name} -> {dest_file.name}")
            filled_count += 1
        except Exception as e:
            log(f"    Failed to copy {missing_id}: {e}", "WARN")

    log(f"  [FILL] Filled {filled_count}/{len(missing_ids)} missing scenes")
    return filled_count, len(missing_ids) - filled_count


# ============================================================================
# PROJECT DETECTION
# ============================================================================

# Folder stability settings
FOLDER_STABLE_CHECK_INTERVAL = 5  # Check every 5 seconds
FOLDER_STABLE_DURATION = 20  # Folder must be stable for 20 seconds


def is_folder_stable(folder_path: Path, check_interval: int = FOLDER_STABLE_CHECK_INTERVAL,
                     stable_duration: int = FOLDER_STABLE_DURATION) -> bool:
    """Check if all files in folder are stable (not being copied/written).

    Returns True if no file sizes have changed for stable_duration seconds.
    This prevents processing a project while VM is still copying files.
    """
    if not folder_path.exists():
        return False

    def get_folder_snapshot():
        """Get dict of {file_path: size} for all files in folder."""
        snapshot = {}
        try:
            for item in folder_path.rglob("*"):
                if item.is_file():
                    try:
                        snapshot[str(item)] = item.stat().st_size
                    except:
                        pass
        except:
            pass
        return snapshot

    checks_needed = stable_duration // check_interval
    last_snapshot = get_folder_snapshot()

    if not last_snapshot:
        return False  # Empty folder

    for i in range(checks_needed):
        time.sleep(check_interval)

        current_snapshot = get_folder_snapshot()

        # Check if any file changed size or new files appeared
        changed = False
        for path, size in current_snapshot.items():
            if path not in last_snapshot:
                # New file appeared
                log(f"    [WAIT] New file detected: {Path(path).name}")
                changed = True
                break
            if last_snapshot[path] != size:
                # File size changed
                log(f"    [WAIT] File still copying: {Path(path).name} ({last_snapshot[path]} -> {size})")
                changed = True
                break

        if changed:
            # Reset and start fresh monitoring
            return is_folder_stable(folder_path, check_interval, stable_duration)

        last_snapshot = current_snapshot

    return True


def get_project_info(project_dir: Path) -> Dict:
    """Get project info from directory."""
    code = project_dir.name

    info = {
        "code": code,
        "path": project_dir,
        "has_srt": False,
        "has_audio": False,
        "has_excel": False,
        "video_count": 0,
        "image_count": 0,
        "media_count": 0,
        "total_scenes": 0,
        "ready_for_edit": False,
        "already_done": False,
    }

    srt_path = project_dir / f"{code}.srt"
    audio_path = project_dir / f"{code}.mp3"
    excel_path = project_dir / f"{code}_prompts.xlsx"

    info["has_srt"] = srt_has_cues(srt_path)
    info["has_audio"] = audio_path.exists()
    info["has_excel"] = excel_path.exists()
    info["srt_path"] = srt_path if info["has_srt"] else None
    info["audio_path"] = audio_path if audio_path.exists() else None
    info["excel_path"] = excel_path if excel_path.exists() else None

    # Support both VM structure (img\ subfolder) and direct structure
    img_dir = project_dir / "img"
    if not img_dir.exists():
        img_dir = project_dir  # Fallback to root folder

    if img_dir.exists():
        videos = [f for f in img_dir.glob("*.mp4")
                  if not f.stem.startswith('nv') and not f.stem.startswith('loc')]
        # Support multiple image formats
        images = []
        for ext in ["*.png", "*.jpg", "*.jpeg", "*.webp"]:
            images.extend([f for f in img_dir.glob(ext)
                          if not f.stem.startswith('nv') and not f.stem.startswith('loc')])
        info["video_count"] = len(videos)
        info["image_count"] = len(images)
        info["media_count"] = len(videos) + len(images)

        if excel_path.exists():
            required_ids = get_required_scene_ids(excel_path)
            info["total_scenes"] = len(required_ids)

    done_dir = DONE_DIR / code
    if done_dir.exists():
        ok, _missing = validate_done_folder(done_dir, require_thumb_folder=False)
        info["already_done"] = ok

    if info["media_count"] > 0 and info["has_audio"] and info["has_excel"]:
        if info["total_scenes"] > 0:
            coverage = info["media_count"] / info["total_scenes"]

            # Fill missing media if coverage >= 10% and < 100%
            if 0.1 <= coverage < 1.0:
                log(f"    - {code}: Coverage {coverage:.0%} < 100%, filling missing...")
                filled, still_missing = fill_missing_media(project_dir, excel_path)

                if filled > 0:
                    videos = [f for f in img_dir.glob("*.mp4")
                              if not f.stem.startswith('nv') and not f.stem.startswith('loc')]
                    images = []
                    for ext in ["*.png", "*.jpg", "*.jpeg", "*.webp"]:
                        images.extend([f for f in img_dir.glob(ext)
                                      if not f.stem.startswith('nv') and not f.stem.startswith('loc')])
                    info["video_count"] = len(videos)
                    info["image_count"] = len(images)
                    info["media_count"] = len(videos) + len(images)
                    coverage = info["media_count"] / info["total_scenes"]
                    log(f"    - {code}: After fill, coverage is now {coverage:.0%}")

            info["ready_for_edit"] = coverage >= 0.5
        else:
            info["ready_for_edit"] = True

    return info


def scan_visual_projects() -> List[Dict]:
    """Scan VISUAL folder for projects ready to edit."""
    projects = []

    if not VISUAL_DIR.exists():
        log(f"VISUAL folder not found: {VISUAL_DIR}", "WARN")
        return projects

    all_folders = [item for item in VISUAL_DIR.iterdir() if item.is_dir()]
    log(f"  [DEBUG] Found {len(all_folders)} folders in VISUAL")

    for item in all_folders:
        info = get_project_info(item)
        code = info["code"]

        # Skip if already being processed (prevent duplicate processing)
        with _processing_lock:
            if code in _processing_codes:
                log(f"    - {code}: SKIPPED (already processing)")
                continue

        if info["already_done"]:
            log(f"    - {code}: already done")
        elif info["ready_for_edit"]:
            log(f"    - {code}: ready ({info['video_count']}v + {info['image_count']}i / {info['total_scenes']} scenes)")
            projects.append(info)
        else:
            reasons = []
            if info["media_count"] == 0:
                reasons.append("no media")
            if not info["has_audio"]:
                reasons.append("no audio")
            if not info["has_excel"]:
                reasons.append("no excel")
            if info["total_scenes"] > 0 and info["media_count"] > 0:
                coverage = info["media_count"] / info["total_scenes"]
                if coverage < 0.5:
                    reasons.append(f"coverage {coverage:.0%} < 50%")
                    # Auto-delete from VISUAL so VMs will redo this project
                    try:
                        visual_path = VISUAL_DIR / code
                        if visual_path.exists():
                            shutil.rmtree(visual_path)
                            log(f"    - {code}: DELETED from VISUAL (coverage {coverage:.0%} < 50%, VMs will redo)")
                    except Exception as e:
                        log(f"    - {code}: Failed to delete: {e}", "WARN")
            log(f"    - {code}: NOT ready ({', '.join(reasons)})")

    return sorted(projects, key=project_priority_key)


def estimate_audio_duration(audio_path: Optional[Path]) -> float:
    """Best-effort audio duration for queue prioritization."""
    if not audio_path or not audio_path.exists():
        return 0.0
    try:
        result = subprocess.run(
            ["ffprobe", "-v", "error", "-show_entries", "format=duration",
             "-of", "default=noprint_wrappers=1:nokey=1", str(audio_path)],
            capture_output=True, text=True, timeout=10, creationflags=SUBPROCESS_FLAGS
        )
        if result.returncode == 0 and result.stdout.strip():
            return max(0.0, float(result.stdout.strip()))
    except Exception:
        pass
    return 0.0


def project_priority_key(project_info: Dict):
    """Shortest-job-first queue order: lighter projects finish sooner."""
    audio_duration = estimate_audio_duration(project_info.get("audio_path"))
    media_count = int(project_info.get("media_count") or 0)
    scene_count = int(project_info.get("total_scenes") or 0)
    video_count = int(project_info.get("video_count") or 0)
    image_count = int(project_info.get("image_count") or 0)
    # Video clips are much heavier than images; voice duration refines ties.
    weight = (video_count * 1.0) + (image_count * 0.35) + (scene_count * 0.15) + (audio_duration / 20.0)
    return (weight, media_count, scene_count, audio_duration, project_info.get("code", ""))


# ============================================================================
# VIDEO COMPOSITION
# ============================================================================

def parse_timestamp(timestamp) -> float:
    """Parse time values like HH:MM:SS,mmm / HH:MM:SS.mmmmmm / MM:SS / seconds."""
    if timestamp is None:
        return 0.0

    if isinstance(timestamp, (int, float)):
        try:
            return max(0.0, float(timestamp))
        except (ValueError, TypeError):
            return 0.0

    ts = str(timestamp).strip()
    if not ts:
        return 0.0

    ts = ts.replace(",", ".")
    frac = 0.0
    if "." in ts:
        ts_part, frac_part = ts.rsplit(".", 1)
        if frac_part.isdigit():
            frac = int(frac_part) / (10 ** len(frac_part))
        else:
            ts_part = ts
    else:
        ts_part = ts

    parts = ts_part.split(":")
    try:
        if len(parts) == 3:
            h, m, s = parts
            base = int(h) * 3600 + int(m) * 60 + int(s)
        elif len(parts) == 2:
            m, s = parts
            base = int(m) * 60 + int(s)
        else:
            return max(0.0, float(ts))
        return max(0.0, base + frac)
    except (ValueError, TypeError):
        return 0.0


def process_srt_for_video(srt_path: Path, output_path: Path, max_chars: int = 45) -> Path:
    """Format SRT for burn without destroying voice-aligned cue timing."""
    MIN_CHUNK_DUR = 0.70

    def to_single_line(text: str) -> str:
        return " ".join(text.split())

    def parse_time(time_str: str) -> float:
        h, m, s = time_str.replace(',', '.').split(':')
        return int(h) * 3600 + int(m) * 60 + float(s)

    def format_time(seconds: float) -> str:
        h = int(seconds // 3600)
        m = int((seconds % 3600) // 60)
        s = seconds % 60
        return f"{h:02d}:{m:02d}:{s:06.3f}".replace('.', ',')

    def choose_break(words: list, max_len: int, target_ratio: float = 0.52) -> int:
        if len(words) <= 1:
            return 1
        glue_prev = {"của", "những", "các", "một", "người", "vì", "để", "khi", "nếu", "rằng", "là"}
        glue_next = {"của", "và", "nhưng", "vì", "để", "khi", "nếu", "rằng", "là", "mà"}
        total_len = len(" ".join(words))
        target = total_len * target_ratio
        best_i = max(1, min(len(words) - 1, len(words) // 2))
        best_score = float("inf")
        for i in range(1, len(words)):
            left = " ".join(words[:i])
            right = " ".join(words[i:])
            prev_word = re.sub(r"\W+$", "", words[i - 1].lower())
            next_word = re.sub(r"^\W+", "", words[i].lower())
            score = abs(len(left) - target)
            score += max(0, len(left) - max_len) * 8
            score += max(0, len(right) - max_len) * 8
            if i <= 2 or len(words) - i <= 2:
                score += 35
            if re.search(r"[.!?…]$", words[i - 1]):
                score -= 45
            elif re.search(r"[,;:]$", words[i - 1]):
                score -= 28
            if prev_word in glue_prev:
                score += 35
            if next_word in glue_next:
                score += 18
            if score < best_score:
                best_score = score
                best_i = i
        return best_i

    def to_display_lines(text: str, max_len: int) -> str:
        # Keep subtitle cues on one line. Fit is handled by font-size profiles.
        return to_single_line(text)

    def split_only_if_too_long(text: str, max_len: int) -> list:
        text = to_single_line(text)
        if len(text) <= max_len * 2:
            return [text]
        chunks = []
        words = text.split()
        while len(" ".join(words)) > max_len * 2 and len(words) > 3:
            i = choose_break(words, max_len * 2, target_ratio=0.50)
            chunks.append(" ".join(words[:i]).strip())
            words = words[i:]
        if words:
            chunks.append(" ".join(words).strip())
        return [c for c in chunks if c]

    try:
        content = read_text_best_effort(srt_path)
        content = unicodedata.normalize("NFC", content).replace("\r\n", "\n").replace("\r", "\n")

        pattern = r'(\d+)\s*\n\s*(\d{2}:\d{2}:\d{2}[,\.]\d{3})\s*-->\s*(\d{2}:\d{2}:\d{2}[,\.]\d{3})\s*\n(.*?)(?=\n\s*\n\s*\d+\s*\n|\Z)'
        entries = re.findall(pattern, content, re.DOTALL)
        if not entries:
            output_path.write_text(content, encoding='utf-8')
            log(f"SRT parse found no cues, kept normalized original: {srt_path.name}", "WARN")
            return output_path if srt_has_cues(output_path) else srt_path

        new_entries = []
        new_index = 1

        for idx, start, end, text in entries:
            text = to_single_line(text.strip().upper())
            start_sec = parse_time(start)
            end_sec = parse_time(end)
            duration = max(0.001, end_sec - start_sec)

            chunks = split_only_if_too_long(text, max_chars)
            if len(chunks) == 1:
                new_entries.append((new_index, start, end, to_display_lines(chunks[0], max_chars)))
                new_index += 1
                continue

            weights = [max(1, len(c.replace("\n", " "))) for c in chunks]
            total_w = sum(weights)
            acc_w = 0
            for i, chunk in enumerate(chunks):
                c_start = start_sec + duration * (acc_w / total_w)
                acc_w += weights[i]
                if i == len(chunks) - 1:
                    c_end = end_sec
                else:
                    c_end = start_sec + duration * (acc_w / total_w)
                if c_end - c_start < MIN_CHUNK_DUR and i < len(chunks) - 1:
                    c_end = min(end_sec, c_start + MIN_CHUNK_DUR)
                new_entries.append((new_index, format_time(c_start), format_time(c_end), to_display_lines(chunk, max_chars)))
                new_index += 1

        with open(output_path, 'w', encoding='utf-8') as f:
            for idx, start, end, text in new_entries:
                f.write(f"{idx}\n{start} --> {end}\n{text}\n\n")

        return output_path

    except Exception as e:
        return srt_path


def shift_srt(srt_path: Path, offset_sec: float, output_path: Path) -> Path:
    """Shift all SRT timestamps by offset_sec (positive = delay sub)."""
    def parse_time(t: str) -> float:
        h, m, s = t.replace(',', '.').split(':')
        return int(h) * 3600 + int(m) * 60 + float(s)

    def fmt_time(s: float) -> str:
        s = max(0.0, s)
        h = int(s // 3600); m = int((s % 3600) // 60); sec = s % 60
        return f"{h:02d}:{m:02d}:{sec:06.3f}".replace('.', ',')

    try:
        text = read_text_best_effort(srt_path)
        def shift_line(m):
            t1 = parse_time(m.group(1)) + offset_sec
            t2 = parse_time(m.group(2)) + offset_sec
            return f"{fmt_time(t1)} --> {fmt_time(t2)}"
        shifted = re.sub(
            r'(\d{2}:\d{2}:\d{2}[,\.]\d{3})\s*-->\s*(\d{2}:\d{2}:\d{2}[,\.]\d{3})',
            shift_line, text
        )
        output_path.write_text(shifted, encoding='utf-8')
        return output_path
    except Exception:
        return srt_path


def shift_srt_with_xfade_compensation(srt_path: Path, media_items: list,
                                      transition_duration: float,
                                      disclaimer_duration: float,
                                      output_path: Path) -> Path:
    """Adjust SRT to visual timeline when using xfade overlap."""
    def parse_time(t: str) -> float:
        h, m, s = t.replace(',', '.').split(':')
        return int(h) * 3600 + int(m) * 60 + float(s)

    def fmt_time(s: float) -> str:
        s = max(0.0, s)
        h = int(s // 3600); m = int((s % 3600) // 60); sec = s % 60
        return f"{h:02d}:{m:02d}:{sec:06.3f}".replace('.', ',')

    try:
        scenes = [item for item in media_items if not item.get('is_disclaimer')]
        scenes_sorted = sorted(scenes, key=lambda x: x.get('start', 0.0))
        scene_mapping = {}
        for i, scene in enumerate(scenes_sorted):
            srt_time = scene.get('start', 0.0)
            actual_time = disclaimer_duration + srt_time - (i * transition_duration)
            scene_mapping[srt_time] = max(0.0, actual_time)

        def map_srt_to_video(srt_t: float) -> float:
            if srt_t <= 0:
                return disclaimer_duration
            scene_times = sorted(scene_mapping.keys())
            if not scene_times:
                return disclaimer_duration + srt_t
            if srt_t <= scene_times[0]:
                first_srt = scene_times[0]
                first_vid = scene_mapping[first_srt]
                ratio = srt_t / first_srt if first_srt > 0 else 0
                return disclaimer_duration + ratio * (first_vid - disclaimer_duration)
            if srt_t >= scene_times[-1]:
                last_srt = scene_times[-1]
                last_vid = scene_mapping[last_srt]
                return srt_t + (last_vid - last_srt)
            for i in range(len(scene_times) - 1):
                t1_srt = scene_times[i]
                t2_srt = scene_times[i + 1]
                if t1_srt <= srt_t <= t2_srt:
                    t1_vid = scene_mapping[t1_srt]
                    t2_vid = scene_mapping[t2_srt]
                    ratio = (srt_t - t1_srt) / (t2_srt - t1_srt) if t2_srt > t1_srt else 0
                    return t1_vid + ratio * (t2_vid - t1_vid)
            return disclaimer_duration + srt_t

        text = read_text_best_effort(srt_path)
        def adjust_line(m):
            t1_srt = parse_time(m.group(1))
            t2_srt = parse_time(m.group(2))
            t1_vid = map_srt_to_video(t1_srt)
            t2_vid = map_srt_to_video(t2_srt)
            return f"{fmt_time(t1_vid)} --> {fmt_time(t2_vid)}"

        adjusted = re.sub(
            r'(\d{2}:\d{2}:\d{2}[,\.]\d{3})\s*-->\s*(\d{2}:\d{2}:\d{2}[,\.]\d{3})',
            adjust_line, text
        )
        output_path.write_text(adjusted, encoding='utf-8')
        return output_path
    except Exception:
        return srt_path


def compose_video(project_info: Dict, callback=None) -> Tuple[bool, Optional[Path], Optional[str]]:
    """Compose final video."""
    code = project_info["code"]
    project_dir = project_info["path"]
    excel_path = project_info.get("excel_path")

    def plog(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            log(f"[{code}] {msg}", level)

    update_progress(code=code, step="Starting", percent=0, status="composing")
    plog("Starting video composition...")
    _t = {'start': time.time()}
    _num_clips = 0

    # Parallel slot info (set by run_scan_loop when running multiple videos)
    _parallel_slot = project_info.get('_slot', 0)       # 0-indexed slot (0, 1, 2...)
    _parallel_count = project_info.get('_parallel', 1)  # total videos in this batch

    if not excel_path or not excel_path.exists():
        return False, None, "Excel file not found"

    # Check FFmpeg
    try:
        result = subprocess.run(["ffmpeg", "-version"], capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS)
        if result.returncode != 0:
            return False, None, "FFmpeg not working"
    except FileNotFoundError:
        return False, None, "FFmpeg not installed"

    # Find voice file
    voice_files = list(project_dir.glob("*.mp3")) + list(project_dir.glob("*.wav"))
    if not voice_files:
        return False, None, "No voice file found"
    voice_path = voice_files[0]

    # Find SRT file
    srt_files = list(project_dir.glob("srt/*.srt")) + list(project_dir.glob("*.srt"))
    srt_path = srt_files[0] if srt_files else None

    output_path = project_dir / f"{code}.mp4"

    # Support both VM structure (img\ subfolder) and direct structure
    img_dir = project_dir / "img"
    if not img_dir.exists():
        # Fallback to project_dir itself if no img\ subfolder
        img_dir = project_dir
        plog(f"  Using root folder for media (no img\\ subfolder)")

    plog(f"  Voice: {voice_path.name}")
    plog(f"  SRT: {srt_path.name if srt_path else 'None'}")
    plog(f"  Excel: {excel_path.name}")

    # Check if Excel file is stable (not being written/downloaded)
    def is_file_stable(file_path, check_interval=5, stable_duration=15):
        """Check if file size is stable for a period of time."""
        if not file_path.exists():
            return False
        last_size = file_path.stat().st_size
        checks = stable_duration // check_interval
        for i in range(checks):
            time.sleep(check_interval)
            if not file_path.exists():
                return False
            current_size = file_path.stat().st_size
            if current_size != last_size:
                plog(f"  Excel still changing: {last_size} -> {current_size}, waiting...")
                last_size = current_size
                return is_file_stable(file_path, check_interval, stable_duration)
        return True

    # Wait for Excel to be stable before opening
    update_progress(code=code, step="Checking Excel", percent=1)
    if not is_file_stable(excel_path):
        return False, None, "Excel file is still being written"

    try:
        import openpyxl
        update_progress(code=code, step="Loading Excel", percent=2)
        wb = openpyxl.load_workbook(excel_path)

        # Find sheet with scene data (must have more than just header row)
        # Priority: sheets with 'scene' in name, then 'srt_coverage' as fallback
        scenes_sheet = None
        fallback_sheet = None
        for sheet_name in wb.sheetnames:
            candidate = wb[sheet_name]
            if candidate.max_row <= 1:
                continue  # Skip empty sheets
            if 'scene' in sheet_name.lower() and sheet_name.lower() != 'srt_coverage':
                scenes_sheet = candidate
                break
            elif sheet_name.lower() == 'srt_coverage':
                fallback_sheet = candidate

        # Use fallback if no primary sheet found
        if not scenes_sheet and fallback_sheet:
            scenes_sheet = fallback_sheet

        if not scenes_sheet:
            return False, None, "No Scenes sheet with data in Excel"

        headers = [cell.value for cell in scenes_sheet[1]]

        id_col = start_col = None
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_lower = str(h).lower().strip()
            if h_lower in ['scene_id', 'id'] and id_col is None:
                id_col = i
            if h_lower == 'srt_start':
                start_col = i
            elif 'start' in h_lower and 'time' in h_lower and start_col is None:
                start_col = i

        if id_col is None:
            return False, None, "No ID column found"

        # Load media
        media_items = []
        video_count = 0
        image_count = 0

        for row in scenes_sheet.iter_rows(min_row=2, values_only=True):
            if row[id_col] is None:
                continue

            scene_id_raw = str(row[id_col]).strip()

            try:
                scene_id_int = int(float(scene_id_raw))
                scene_id = str(scene_id_int)
            except ValueError:
                continue

            media_path = None
            is_video = False

            possible_ids = [scene_id, f"{scene_id}.0", scene_id_raw]
            possible_ids = list(dict.fromkeys(possible_ids))

            for sid in possible_ids:
                video_path = img_dir / f"{sid}.mp4"

                if video_path.exists():
                    media_path = video_path
                    is_video = True
                    video_count += 1
                    break

                # Check multiple image formats
                for img_ext in [".png", ".jpg", ".jpeg", ".webp"]:
                    img_path = img_dir / f"{sid}{img_ext}"
                    if img_path.exists():
                        media_path = img_path
                        is_video = False
                        image_count += 1
                        break
                if media_path:
                    break

            if not media_path:
                continue

            start_time = 0.0
            if start_col is not None and row[start_col]:
                start_time = parse_timestamp(str(row[start_col]))

            media_items.append({
                'id': scene_id,
                'path': str(media_path),
                'start': start_time,
                'is_video': is_video
            })

        if not media_items:
            return False, None, "No media found in img/ folder"

        media_items.sort(key=lambda x: x['start'])
        plog(f"  Found {len(media_items)} media: {video_count} videos, {image_count} images")
        update_progress(code=code, step=f"Found {len(media_items)} media", percent=3)

        # Handle gap at start
        GAP_THRESHOLD = 0.5
        first_start = media_items[0]['start']

        if first_start > GAP_THRESHOLD:
            plog(f"  Gap at start: 0:00 -> {first_start:.1f}s, using first media as filler")
            filler_item = {
                'id': f"{media_items[0]['id']}_filler",
                'path': media_items[0]['path'],
                'start': 0.0,
                'is_video': media_items[0]['is_video'],
                'is_filler': True
            }
            media_items.insert(0, filler_item)

        # Get voice duration
        probe_cmd = ["ffprobe", "-v", "error", "-show_entries", "format=duration",
                    "-of", "default=noprint_wrappers=1:nokey=1", str(voice_path)]
        result = subprocess.run(probe_cmd, capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS)
        total_duration = float(result.stdout.strip()) if result.stdout.strip() else 60.0
        plog(f"  Voice duration: {total_duration:.1f}s")

        # Calculate durations
        for i, item in enumerate(media_items):
            if i == 0:
                if len(media_items) > 1:
                    item['duration'] = media_items[1]['start']
                else:
                    item['duration'] = total_duration
            elif i < len(media_items) - 1:
                item['duration'] = media_items[i + 1]['start'] - item['start']
            else:
                item['duration'] = total_duration - item['start']

            if item['duration'] <= 0:
                item['duration'] = max(0.5, (total_duration - item['start']) / max(1, len(media_items) - i))

        # Create video
        temp_dir = tempfile.mkdtemp()
        try:
            temp_video = Path(temp_dir) / "temp_video.mp4"

            # Normalize voice before timing/mixing so merged ElevenLabs chunks stay even.
            update_progress(code=code, step="Normalizing audio", percent=3)
            plog("  Mastering voice (-14 LUFS, TP -1.5 dB, local leveling)...")
            norm_voice = normalize_voice(voice_path, Path(temp_dir), label="voice")
            if norm_voice != voice_path:
                plog(f"  Voice normalized: {voice_path.name} -> {norm_voice.name}")
                voice_path = norm_voice
            else:
                plog("  Voice normalize fallback: using original file", "WARN")

            # Load settings from channel template (per-channel customization)
            update_progress(code=code, step="Loading settings", percent=4)
            channel_template = get_subtitle_template(code)
            channel = code.split("-")[0] if "-" in code else code

            # --- Disclaimer image (1s static at video start) ---
            DISCLAIMER_DURATION = 1.0
            IMAGES_DIR = TOOL_DIR / "images"
            disclaimer_img = None
            audio_delay_offset = 0.0  # Track audio delay needed for disclaimer sync
            if IMAGES_DIR.exists() and media_items and media_items[0].get('duration', 0) > DISCLAIMER_DURATION + 0.5:
                # Try channel-specific image: KA2-T2.jpg, KA2-T*.jpg, KA2.jpg
                for pattern in [f"{channel}-T*.jpg", f"{channel}-T*.png",
                                 f"{channel}.jpg", f"{channel}.png"]:
                    matches = sorted(IMAGES_DIR.glob(pattern))
                    if matches:
                        disclaimer_img = matches[0]
                        break

            if disclaimer_img:
                plog(f"  Disclaimer: {disclaimer_img.name} ({DISCLAIMER_DURATION}s) prepended to video")
                # Prepend disclaimer WITHOUT subtracting from scene 1.
                # Scene 1 keeps its full srt_dur. Timing is handled via
                # xfade padding (+T per clip) and abs_desired offsets.
                media_items.insert(0, {
                    'id': 'disclaimer',
                    'path': str(disclaimer_img),
                    'start': -DISCLAIMER_DURATION,
                    'duration': DISCLAIMER_DURATION,
                    'is_video': False,
                    'is_disclaimer': True,
                })

            # Video settings from template (with defaults)
            output_resolution = channel_template.get("output_resolution", "4k").lower()
            compose_mode = channel_template.get("compose_mode", "quality").lower()
            kb_intensity = channel_template.get("ken_burns_intensity", "subtle").lower()
            video_transition = channel_template.get("video_transition", "random").lower()
            output_fps = 30  # Fixed FPS
            transition_duration = 0.5
            veo_crop_enabled = _to_bool(channel_template.get("veo_crop_enabled", True), True)
            veo_crop_right_ratio = _to_float(channel_template.get("veo_crop_right_ratio", 0.05), 0.05)
            veo_crop_bottom_ratio = _to_float(channel_template.get("veo_crop_bottom_ratio", 0.07), 0.07)
            veo_crop_min_right_px = _to_int(channel_template.get("veo_crop_min_right_px", 56), 56)
            veo_crop_min_bottom_px = _to_int(channel_template.get("veo_crop_min_bottom_px", 48), 48)
            veo_crop_keep_4k_aspect = _to_bool(channel_template.get("veo_crop_keep_4k_aspect", True), True)
            veo_crop_right_ratio = max(0.0, min(0.2, veo_crop_right_ratio))
            veo_crop_bottom_ratio = max(0.0, min(0.2, veo_crop_bottom_ratio))
            veo_crop_min_right_px = max(2, min(200, veo_crop_min_right_px))
            veo_crop_min_bottom_px = max(2, min(200, veo_crop_min_bottom_px))

            # Fallback to global config if no template settings
            try:
                import yaml
                config_path = TOOL_DIR / "config" / "settings.yaml"
                if config_path.exists():
                    with open(config_path, 'r', encoding='utf-8') as f:
                        config = yaml.safe_load(f) or {}
                    # Only use global config if not set in template
                    if "output_resolution" not in channel_template:
                        output_resolution = config.get('output_resolution', '4k').lower()
                    if "compose_mode" not in channel_template:
                        compose_mode = config.get('video_compose_mode', 'quality').lower()
                    output_fps = config.get('output_fps', 30)
                    transition_duration = config.get('transition_duration', 0.5)
            except:
                pass

            # Get prefer_gpu setting from template (default: auto)
            # "auto" = use GPU when available and compose_mode is not "quality"
            # "always" = always prefer GPU over OpenCV (faster but slightly lower quality)
            # "never" = always use OpenCV when available
            prefer_gpu = channel_template.get("prefer_gpu", "auto").lower()

            plog(f"  Channel: {channel} | Res: {output_resolution.upper()} | Mode: {compose_mode}")
            if veo_crop_enabled:
                plog(
                    f"  Veo crop: ON (right={veo_crop_right_ratio:.3f}, bottom={veo_crop_bottom_ratio:.3f}, "
                    f"min={veo_crop_min_right_px}px/{veo_crop_min_bottom_px}px, keep_aspect={veo_crop_keep_4k_aspect})"
                )
            else:
                plog("  Veo crop: OFF")

            # Determine if using xfade transitions (all presets except "none" use xfade)
            use_xfade = video_transition != "none"
            FADE_DURATION = transition_duration if use_xfade else 0.4

            # FIX: xfade timing compensation via padding.
            # The xfade filter overlaps adjacent clips by T seconds, so each clip
            # is consumed T seconds earlier than its duration. By padding each
            # non-last clip with +T, the effective clip duration matches srt_dur.
            # Combined with abs_desired offsets below, this places each scene at
            # exactly its srt_start in the video.
            if use_xfade:
                for i, item in enumerate(media_items):
                    if i < len(media_items) - 1:  # All clips except last get +T
                        item['duration'] = item['duration'] + FADE_DURATION
                plog(f"  xfade duration pad: +{FADE_DURATION}s per clip")

            # Use cached system resources (detected once at startup)
            resources = get_system_resources()
            use_gpu = resources.get("gpu_available", False)
            gpu_encoder = resources.get("gpu_encoder", "libx264")

            # Slot 0 = GPU, Slot 1+ = CPU (unless VE3_GPU_SLOTS allows more)
            # VE3_GPU_SLOTS=2 means slot 0 and 1 both use GPU (for 12GB+ VRAM cards)
            gpu_slots = int(os.environ.get("VE3_GPU_SLOTS", 1))
            if _parallel_slot >= gpu_slots and use_gpu:
                use_gpu = False
                gpu_encoder = "libx264"
                plog(f"  Slot {_parallel_slot}: CPU mode (GPU slots: 0-{gpu_slots-1})")

            if use_gpu:
                plog(f"  GPU Encoder: {gpu_encoder.upper()}")

            # Determine whether to use OpenCV Ken Burns or FFmpeg
            # OpenCV = higher quality but CPU intensive
            # FFmpeg+GPU = faster but slightly lower Ken Burns quality
            if prefer_gpu == "always" and use_gpu:
                # Force GPU mode - skip OpenCV
                use_opencv_kb = False
                plog(f"  Mode: GPU-accelerated (prefer_gpu=always)")
            elif prefer_gpu == "never":
                # Force OpenCV mode
                use_opencv_kb = KEN_BURNS_CV2_AVAILABLE and compose_mode in ["quality", "balanced"]
            else:  # auto
                # Default: use OpenCV for quality/balanced, GPU for fast
                if compose_mode == "fast" or not KEN_BURNS_CV2_AVAILABLE:
                    use_opencv_kb = False
                else:
                    use_opencv_kb = compose_mode in ["quality", "balanced"]

            # Initialize Ken Burns generator
            # For xfade transitions, don't apply individual clip fades (xfade handles it)
            clip_fade_duration = 0.0 if use_xfade else FADE_DURATION

            # OPTIMIZATION: Render at 1080p internally, upscale to final resolution later
            # This reduces processing time by ~4x for 4K output
            final_output_resolution = output_resolution
            final_output_size = QUALITY_PRESETS.get(output_resolution, QUALITY_PRESETS["1080p"])
            render_resolution = "1080p"  # Always render at 1080p for speed
            render_size = QUALITY_PRESETS["1080p"]
            needs_upscale = output_resolution in ["4k", "2k"]

            if use_opencv_kb:
                ken_burns = KenBurnsCv2(
                    output_resolution=render_resolution,  # Render at 1080p
                    fps=output_fps,
                    fade_duration=clip_fade_duration,
                    intensity=kb_intensity
                )
                plog(f"  Ken Burns intensity: {kb_intensity.upper()}")
                # Determine output size
                if output_resolution == "auto":
                    # Use first media to detect
                    first_media = media_items[0]['path']
                    import cv2
                    if media_items[0]['is_video']:
                        cap = cv2.VideoCapture(str(first_media))
                        w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                        h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                        cap.release()
                    else:
                        img = cv2.imread(str(first_media))
                        h, w = img.shape[:2]
                    final_output_size = ken_burns.detect_optimal_resolution(w, h)
                    # Still render at lower res for speed
                    if final_output_size[0] >= 3840:
                        render_size = QUALITY_PRESETS["1080p"]
                        needs_upscale = True
                    elif final_output_size[0] >= 2560:
                        render_size = QUALITY_PRESETS["1080p"]
                        needs_upscale = True
                    else:
                        render_size = final_output_size
                        needs_upscale = False

                output_size = render_size  # Use render size for clip creation
                plog(f"  Render: {render_size[0]}x{render_size[1]} -> Final: {final_output_size[0]}x{final_output_size[1]} @ {output_fps}fps")
            else:
                # Fallback to FFmpeg Ken Burns
                ken_burns = KenBurnsGenerator(1920, 1080, intensity="normal")
                output_size = (1920, 1080)

            # Reduce clip workers when running parallel to avoid CPU overload
            # Solo: 8 workers. Parallel 2x: 5 workers each (5+5=10 on 8 cores, better than 8+8=16)
            _reduce = int(os.environ.get("VE3_WORKER_REDUCE", 3))
            clip_workers = max(4, CLIP_WORKERS - _reduce) if _parallel_count >= 2 else CLIP_WORKERS

            plog(f"  Compose mode: {compose_mode.upper()} ({'OpenCV' if use_opencv_kb else 'FFmpeg'})")
            plog(f"  Transition: {video_transition.upper()} ({transition_duration}s)")
            if _parallel_count >= 2:
                sub_mode = "NVENC" if use_gpu else "CPU"
                plog(f"  Parallel: slot {_parallel_slot}/{_parallel_count-1} | Clips: {clip_workers} workers | Sub: {sub_mode}")
            plog(f"  Creating {len(media_items)} clips with {clip_workers} parallel workers...")
            total_clips = len(media_items)
            update_progress(code=code, step="Creating clips", percent=5, clip_total=total_clips)

            # Helper function for creating a single clip
            def create_single_clip(task_args):
                """Create a single clip - runs in parallel worker."""
                (idx, item_data, clip_path_str, kb_config) = task_args
                clip_path = Path(clip_path_str)
                media_path = Path(item_data['path'])
                target_duration = item_data['duration']
                is_video = item_data['is_video']
                success = False

                # Disclaimer: static image, no Ken Burns, no fade
                if item_data.get('is_disclaimer'):
                    abs_path = str(media_path.resolve()).replace('\\', '/')
                    out_w, out_h = kb_config['output_size']
                    vf = f"scale={out_w}:{out_h}:force_original_aspect_ratio=decrease,pad={out_w}:{out_h}:(ow-iw)/2:(oh-ih)/2"
                    if kb_config['use_gpu']:
                        cmd_d = ["ffmpeg", "-y", "-loop", "1", "-t", str(target_duration),
                                 "-i", abs_path, "-vf", vf, "-c:v", kb_config['gpu_encoder'],
                                 "-preset", "p5", "-rc", "vbr", "-cq", "22",
                                 "-pix_fmt", "yuv420p", "-r", str(kb_config['fps']), str(clip_path)]
                    else:
                        cmd_d = ["ffmpeg", "-y", "-loop", "1", "-t", str(target_duration),
                                 "-i", abs_path, "-vf", vf, "-c:v", "libx264",
                                 "-preset", "fast", "-pix_fmt", "yuv420p",
                                 "-r", str(kb_config['fps']), str(clip_path)]
                    res_d = subprocess.run(cmd_d, capture_output=True, text=True, timeout=60, creationflags=SUBPROCESS_FLAGS)
                    return (idx, res_d.returncode == 0,
                            str(clip_path) if res_d.returncode == 0 and clip_path.exists() else None)

                # Pre-validate image file (skip corrupted images)
                if not is_video:
                    try:
                        from PIL import Image
                        with Image.open(media_path) as test_img:
                            test_img.load()  # Force load to detect corruption
                    except Exception as e:
                        print(f"    Cannot read image: {media_path}")
                        return (idx, False, None)  # Skip this image

                # Create worker-local Ken Burns instance for thread safety
                worker_kb = None
                if kb_config['use_opencv']:
                    try:
                        worker_kb = KenBurnsCv2(
                            output_resolution=kb_config['resolution'],
                            fps=kb_config['fps'],
                            fade_duration=kb_config['fade_duration'],
                            intensity=kb_config['intensity']
                        )
                    except:
                        pass

                if is_video:
                    # Process VIDEO clip via FFmpeg (more reliable than OpenCV for trimming)
                    # OpenCV has duration accuracy issues with mp4v container
                    success = False  # Always use FFmpeg for videos

                    if not success:
                        # Fallback to FFmpeg
                        abs_path = str(media_path.resolve()).replace('\\', '/')
                        probe_cmd = [
                            "ffprobe", "-v", "error",
                            "-show_entries", "format=duration:stream=width,height",
                            "-select_streams", "v:0",
                            "-of", "json", abs_path
                        ]
                        probe_result = subprocess.run(probe_cmd, capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS)
                        video_duration = 8.0
                        src_w, src_h = 0, 0
                        if probe_result.returncode == 0 and probe_result.stdout.strip():
                            try:
                                probe_data = json.loads(probe_result.stdout)
                                video_duration = float((probe_data.get("format", {}) or {}).get("duration") or 8.0)
                                streams = probe_data.get("streams") or []
                                if streams:
                                    src_w = int(streams[0].get("width") or 0)
                                    src_h = int(streams[0].get("height") or 0)
                            except Exception:
                                pass

                        out_w, out_h = kb_config['output_size']
                        crop_prefix = ""
                        crop_info = None
                        if kb_config.get('veo_crop_enabled', True) and src_w > 0 and src_h > 0:
                            crop_info = compute_veo_crop_geometry(
                                src_w=src_w,
                                src_h=src_h,
                                target_w=out_w,
                                target_h=out_h,
                                right_ratio=kb_config.get('veo_crop_right_ratio', 0.05),
                                bottom_ratio=kb_config.get('veo_crop_bottom_ratio', 0.07),
                                min_right_px=kb_config.get('veo_crop_min_right_px', 56),
                                min_bottom_px=kb_config.get('veo_crop_min_bottom_px', 48),
                                keep_target_aspect=kb_config.get('veo_crop_keep_4k_aspect', True),
                            )
                        if crop_info:
                            crop_prefix = f"crop={crop_info['crop_w']}:{crop_info['crop_h']}:{crop_info['crop_x']}:{crop_info['crop_y']},"
                            if idx < 3:
                                print(
                                    f"    Veo crop applied [{media_path.name}]: "
                                    f"right={crop_info['right_crop']}px bottom={crop_info['bottom_crop']}px "
                                    f"({src_w}x{src_h} -> {crop_info['crop_w']}x{crop_info['crop_h']})"
                                )

                        if crop_info:
                            # Crop already aligned to target aspect -> direct scale, no pad/letterbox.
                            base_vf = f"{crop_prefix}scale={out_w}:{out_h}:flags=lanczos,setsar=1"
                        else:
                            # Fallback: fill frame without black bars.
                            base_vf = f"scale={out_w}:{out_h}:force_original_aspect_ratio=increase,crop={out_w}:{out_h},setsar=1"
                        if kb_config['use_xfade']:
                            vf = base_vf
                        else:
                            vf = f"{base_vf},fade=t=in:st=0:d={kb_config['fade_dur']},fade=t=out:st={max(0, target_duration - kb_config['fade_dur'])}:d={kb_config['fade_dur']}"

                        v_encoder = kb_config['gpu_encoder'] if kb_config['use_gpu'] else "libx264"
                        v_preset = ["-preset", "p4"] if kb_config['use_gpu'] else ["-preset", "medium"]

                        if video_duration > target_duration:
                            # Video longer → trim from center
                            trim_start = (video_duration - target_duration) / 2
                            cmd_clip = [
                                "ffmpeg", "-y", "-ss", str(trim_start), "-i", abs_path,
                                "-t", str(target_duration), "-vf", vf,
                                "-c:v", v_encoder, *v_preset, "-pix_fmt", "yuv420p",
                                "-an", "-r", str(kb_config['fps']), str(clip_path)
                            ]
                        elif video_duration < target_duration and video_duration > 0.5:
                            # Video shorter → loop to exceed target, then trim to exact duration
                            import math
                            loop_count = math.ceil(target_duration / video_duration)
                            # Add small buffer to ensure smooth loop (extra 0.1x)
                            loop_count = max(2, loop_count)
                            looped_duration = video_duration * loop_count

                            # Use concat demuxer for seamless loop (better than -stream_loop for accuracy)
                            # Create looped video, then trim to exact target_duration
                            loop_vf = f"{vf}"
                            cmd_clip = [
                                "ffmpeg", "-y",
                                "-stream_loop", str(loop_count - 1), "-i", abs_path,
                                "-t", str(target_duration),
                                "-vf", loop_vf,
                                "-c:v", v_encoder, *v_preset,
                                "-pix_fmt", "yuv420p", "-an", "-r", str(kb_config['fps']),
                                str(clip_path)
                            ]
                            if idx < 3:
                                print(f"    Loop video [{media_path.name}]: {video_duration:.1f}s × {loop_count} = {looped_duration:.1f}s → trim to {target_duration:.1f}s")
                        else:
                            # Video same duration → use as-is
                            cmd_clip = [
                                "ffmpeg", "-y", "-i", abs_path, "-t", str(target_duration),
                                "-vf", vf, "-c:v", v_encoder, *v_preset,
                                "-pix_fmt", "yuv420p", "-an", "-r", str(kb_config['fps']), str(clip_path)
                            ]

                        result = subprocess.run(cmd_clip, capture_output=True, text=True, timeout=300, creationflags=SUBPROCESS_FLAGS)
                        success = result.returncode == 0
                else:
                    # Process IMAGE with Ken Burns effect
                    if worker_kb:
                        success = worker_kb.create_clip_from_image(
                            media_path, clip_path, target_duration,
                            effect=None,
                            output_size=kb_config['output_size']
                        )

                    if not success:
                        # Fallback to FFmpeg
                        abs_path = str(media_path.resolve()).replace('\\', '/')
                        out_w, out_h = kb_config['output_size']
                        base_filter = f"scale={out_w}:{out_h}:force_original_aspect_ratio=decrease,pad={out_w}:{out_h}:(ow-iw)/2:(oh-ih)/2"
                        if kb_config['use_xfade']:
                            vf = base_filter
                        else:
                            vf = f"{base_filter},fade=t=in:st=0:d={kb_config['fade_dur']},fade=t=out:st={max(0, target_duration - kb_config['fade_dur'])}:d={kb_config['fade_dur']}"

                        if kb_config['use_gpu']:
                            cmd_clip = [
                                "ffmpeg", "-y", "-loop", "1", "-t", str(target_duration),
                                "-i", abs_path, "-vf", vf, "-c:v", kb_config['gpu_encoder'],
                                "-preset", "p5", "-rc", "vbr", "-cq", "22",
                                "-pix_fmt", "yuv420p", "-r", str(kb_config['fps']), str(clip_path)
                            ]
                        else:
                            cpu_preset = "ultrafast" if kb_config['compose_mode'] == "fast" else "medium"
                            cmd_clip = [
                                "ffmpeg", "-y", "-loop", "1", "-t", str(target_duration),
                                "-i", abs_path, "-vf", vf, "-c:v", "libx264",
                                "-preset", cpu_preset, "-pix_fmt", "yuv420p", "-r", str(kb_config['fps']), str(clip_path)
                            ]

                        result = subprocess.run(cmd_clip, capture_output=True, text=True, timeout=300, creationflags=SUBPROCESS_FLAGS)
                        success = result.returncode == 0

                return (idx, success, str(clip_path) if success and clip_path.exists() else None)

            # Prepare config for workers
            # Use render_resolution (1080p) for clip creation, upscale to final later
            kb_config = {
                'use_opencv': use_opencv_kb,
                'resolution': render_resolution,  # Render at 1080p for speed
                'fps': output_fps,
                'fade_duration': clip_fade_duration,
                'intensity': kb_intensity,
                'output_size': render_size,  # Use render size (1080p)
                'use_xfade': use_xfade,
                'fade_dur': FADE_DURATION,
                'use_gpu': use_gpu,
                'gpu_encoder': gpu_encoder,
                'compose_mode': compose_mode,
                'veo_crop_enabled': veo_crop_enabled,
                'veo_crop_right_ratio': veo_crop_right_ratio,
                'veo_crop_bottom_ratio': veo_crop_bottom_ratio,
                'veo_crop_min_right_px': veo_crop_min_right_px,
                'veo_crop_min_bottom_px': veo_crop_min_bottom_px,
                'veo_crop_keep_4k_aspect': veo_crop_keep_4k_aspect,
            }

            # Create task list
            clip_tasks = []
            for i, item in enumerate(media_items):
                clip_path = Path(temp_dir) / f"clip_{i:03d}.mp4"
                clip_tasks.append((i, item, str(clip_path), kb_config))

            # Process clips in parallel
            clip_results = [None] * total_clips
            completed_count = 0

            with ThreadPoolExecutor(max_workers=clip_workers) as executor:
                futures = {executor.submit(create_single_clip, task): task[0] for task in clip_tasks}

                for future in as_completed(futures):
                    idx, success, clip_path_str = future.result()
                    if success and clip_path_str:
                        clip_results[idx] = Path(clip_path_str)

                    completed_count += 1
                    clip_percent = 5 + int(completed_count / total_clips * 65)
                    update_progress(code=code, clip_current=completed_count, percent=clip_percent)

                    if completed_count % 20 == 0:
                        plog(f"  ... {completed_count}/{total_clips} clips")

            # Collect successful clips in order, keeping original media_items index for timing
            # clip_index_map[clip_position] = media_items_index (needed for xfade_srt_offsets lookup)
            clip_paths = []
            clip_media_indices = []  # parallel list: which media_items[i] each clip belongs to
            for orig_idx, p in enumerate(clip_results):
                if p is not None:
                    clip_paths.append(p)
                    clip_media_indices.append(orig_idx)

            if not clip_paths:
                return False, None, "No clips created"

            plog(f"  Created {len(clip_paths)} clips...")
            _t['clips'] = time.time()
            _num_clips = len(clip_paths)

            def _probe_clip_duration(cp):
                try:
                    if not cp.exists() or cp.stat().st_size <= 1024:
                        return 0.0
                    probe_cmd = ["ffprobe", "-v", "error", "-show_entries", "format=duration",
                                 "-of", "default=noprint_wrappers=1:nokey=1", str(cp)]
                    probe_result = subprocess.run(
                        probe_cmd,
                        capture_output=True,
                        text=True,
                        timeout=20,
                        creationflags=SUBPROCESS_FLAGS,
                    )
                    if probe_result.returncode != 0:
                        return 0.0
                    return float(probe_result.stdout.strip() or 0.0)
                except Exception:
                    return 0.0

            def _clip_is_valid(cp, min_duration=0.05):
                return _probe_clip_duration(cp) >= min_duration

            def _repair_clip_list(paths, label):
                repaired = []
                bad = []
                for clip_idx, clip_path in enumerate(paths):
                    clip_path = Path(clip_path)
                    if _clip_is_valid(clip_path):
                        repaired.append(clip_path)
                        continue
                    fallback_clip = Path(temp_dir) / f"clip_{clip_idx:03d}.mp4"
                    if fallback_clip != clip_path and _clip_is_valid(fallback_clip):
                        plog(f"    {label}: {clip_path.name} invalid, using {fallback_clip.name}", "WARN")
                        repaired.append(fallback_clip)
                    else:
                        bad.append(clip_path.name)
                        repaired.append(clip_path)
                if bad:
                    plog(f"    {label}: invalid clips remain: {', '.join(bad[:8])}", "WARN")
                return repaired, bad

            # Re-encode clips for high quality xfade (avoid artifacts during crossfade)
            # Also upscale to final resolution if needed (1080p -> 4K)
            if use_xfade and len(clip_paths) > 1:
                upscale_msg = f" + upscale to {final_output_size[0]}x{final_output_size[1]}" if needs_upscale else ""
                plog(f"  Re-encoding {len(clip_paths)} clips for smooth crossfade{upscale_msg} (parallel)...")
                update_progress(code=code, step="Optimizing clips", percent=72)

                def reencode_single_clip(args):
                    """Re-encode a single clip - runs in parallel."""
                    i, cp, temp_dir_path, use_gpu_enc, gpu_enc, do_upscale, final_size, target_dur = args
                    reenc_path = Path(temp_dir_path) / f"hq_{i:03d}.mp4"
                    try:
                        if reenc_path.exists():
                            reenc_path.unlink()
                    except:
                        pass

                    # Build scale filter for upscaling
                    scale_filter = f"scale={final_size[0]}:{final_size[1]}:flags=lanczos" if do_upscale else ""

                    # -t ensures reencoded clip is exactly target duration (prevents FFmpeg frame rounding drift)
                    t_arg = ["-t", f"{target_dur:.6f}"] if target_dur and target_dur > 0 else []

                    if use_gpu_enc:
                        base_cmd = [
                            "ffmpeg", "-y", "-i", str(cp),
                        ]
                        if scale_filter:
                            base_cmd.extend(["-vf", scale_filter])
                        base_cmd.extend(t_arg + [
                            "-c:v", gpu_enc, "-preset", "p5",
                            "-rc", "vbr", "-cq", "18", "-b:v", "25M",
                            "-pix_fmt", "yuv420p", "-r", "30", "-an",
                            str(reenc_path)
                        ])
                        reencode_cmd = base_cmd
                    else:
                        base_cmd = [
                            "ffmpeg", "-y", "-i", str(cp),
                        ]
                        if scale_filter:
                            base_cmd.extend(["-vf", scale_filter])
                        base_cmd.extend(t_arg + [
                            "-c:v", "libx264", "-preset", "fast", "-crf", "17",
                            "-profile:v", "high", "-pix_fmt", "yuv420p", "-r", "30", "-an",
                            str(reenc_path)
                        ])
                        reencode_cmd = base_cmd

                    try:
                        result = subprocess.run(reencode_cmd, capture_output=True, text=True, timeout=180, creationflags=SUBPROCESS_FLAGS)
                        reenc_ok = (result.returncode == 0 and _clip_is_valid(reenc_path))
                        if result.returncode == 0 and reenc_ok:
                            # Keep original clip until final cleanup so we always have a fallback.
                            return (i, reenc_path)
                        else:
                            try:
                                if reenc_path.exists() and not _clip_is_valid(reenc_path):
                                    reenc_path.unlink()
                            except:
                                pass
                            return (i, cp)  # Keep original if re-encode fails
                    except:
                        try:
                            if reenc_path.exists() and not _clip_is_valid(reenc_path):
                                reenc_path.unlink()
                        except:
                            pass
                        return (i, cp)

                # Keep clips at render_size (1080p) for xfade - upscale happens after merge in subtitle burn
                # Pass target_duration so reencode trims exactly to correct length (prevents frame rounding drift)
                reencode_tasks = [
                    (i, cp, temp_dir, use_gpu, gpu_encoder, False, render_size,
                     media_items[clip_media_indices[i]]['duration'])
                    for i, cp in enumerate(clip_paths)
                ]

                # Use parallel workers (limit for GPU VRAM: each ~500MB)
                reencode_cap = int(os.environ.get("VE3_REENCODE_WORKERS", 6))
                reencode_workers = min(reencode_cap, clip_workers) if use_gpu else clip_workers
                reencoded_results = []

                with ThreadPoolExecutor(max_workers=reencode_workers) as executor:
                    futures = {executor.submit(reencode_single_clip, task): task[0] for task in reencode_tasks}
                    completed = 0
                    for future in as_completed(futures):
                        result = future.result()
                        reencoded_results.append(result)
                        completed += 1
                        if completed % 10 == 0:
                            update_progress(code=code, step="Optimizing clips", clip_current=completed, clip_total=len(clip_paths))

                # Sort by index to maintain order
                reencoded_results.sort(key=lambda x: x[0])
                repaired_clip_paths = []
                for clip_idx, clip_path in reencoded_results:
                    clip_ok = _clip_is_valid(clip_path)
                    if not clip_ok:
                        fallback_clip = Path(temp_dir) / f"clip_{clip_idx:03d}.mp4"
                        if _clip_is_valid(fallback_clip):
                            plog(f"    Re-encode output invalid, using original clip_{clip_idx:03d}.mp4", "WARN")
                            clip_path = fallback_clip
                            clip_ok = True
                    if not clip_ok:
                        plog(f"    Clip invalid after re-encode: {clip_path.name}", "WARN")
                    repaired_clip_paths.append(clip_path)
                clip_paths = repaired_clip_paths
                plog(f"  Re-encoded {len(clip_paths)} clips")
            _t['reencode'] = time.time()

            clip_paths, invalid_after_reencode = _repair_clip_list(clip_paths, "Pre-concat validation")
            if invalid_after_reencode:
                return False, None, f"Invalid clips before concat: {', '.join(invalid_after_reencode[:8])}"

            update_progress(code=code, step="Concatenating", percent=75)

            # Concat with appropriate transition
            if use_xfade and len(clip_paths) > 1:
                # Use xfade filter for smooth transitions
                # Build xfade filter chain
                def get_xfade_type(transition_setting):
                    """
                    Get xfade transition type.
                    Presets:
                      mix       - dissolve only (smoothest)
                      cinematic - fade/zoom/circle, documentary style
                      dynamic   - slide/cover/reveal, energetic
                      soft      - gentle fades only
                      news      - wipe styles, professional
                      wipe      - directional wipes
                      random    - weighted mix of all beautiful transitions
                    """
                    if transition_setting == "mix":
                        return "dissolve"

                    elif transition_setting == "cinematic":
                        return random.choice([
                            "fade", "dissolve", "fadeblack",
                            "fadeslow", "fadegrays", "distance",
                            "zoomin", "circleopen",
                        ])

                    elif transition_setting == "dynamic":
                        return random.choice([
                            "slideleft", "slideright", "slideup", "slidedown",
                            "coverleft", "coverright", "coverup", "coverdown",
                            "revealleft", "revealright",
                            "zoomin", "pixelize",
                        ])

                    elif transition_setting == "soft":
                        return random.choice([
                            "fade", "dissolve", "fadeblack", "fadewhite",
                            "fadeslow", "fadefast", "fadegrays",
                        ])

                    elif transition_setting == "news":
                        return random.choice([
                            "wipeleft", "wiperight", "wipeup", "wipedown",
                            "wipetl", "wipetr", "wipebl", "wipebr",
                            "horzopen", "radial",
                        ])

                    elif transition_setting == "wipe":
                        return random.choice([
                            "wipeleft", "wiperight", "wipeup", "wipedown",
                        ])

                    elif transition_setting == "random":
                        # Weighted: smooth fades most common, variety added
                        r = random.random()
                        if r < 0.30:
                            return random.choice(["fade", "dissolve", "fadeblack", "fadegrays"])
                        elif r < 0.50:
                            return random.choice(["fadeslow", "fadefast", "zoomin", "distance"])
                        elif r < 0.70:
                            return random.choice([
                                "coverleft", "coverright", "coverup", "coverdown",
                                "revealleft", "revealright",
                            ])
                        elif r < 0.85:
                            return random.choice([
                                "slideleft", "slideright", "slideup", "slidedown",
                            ])
                        else:
                            return random.choice([
                                "circleopen", "horzopen", "radial",
                                "pixelize", "squeezeh", "squeezev",
                                "wipeleft", "wiperight",
                            ])

                    return "dissolve"


                def xfade_batch(batch_clips, batch_idx, temp_dir_path, batch_start_index=0):
                    """Process a batch of clips with xfade transitions"""
                    if len(batch_clips) == 1:
                        return batch_clips[0]  # Single clip, no xfade needed

                    # Get clip durations + validate each clip
                    batch_durations = []
                    for cp in batch_clips:
                        # Validate file exists and has content
                        dur = _probe_clip_duration(cp)
                        if dur <= 0:
                            size = cp.stat().st_size if cp.exists() else 0
                            plog(f"    Clip invalid/missing: {cp.name} ({size} bytes)", "WARN")
                            return None
                        if dur < transition_duration:
                            plog(f"    Clip too short for xfade: {cp.name} ({dur:.2f}s < {transition_duration}s)", "WARN")
                            return None
                        batch_durations.append(dur)

                    # Build inputs
                    inputs = []
                    for cp in batch_clips:
                        inputs.extend(["-i", str(cp).replace('\\', '/')])

                    # Build filter_complex using srt_start-based offsets.
                    # Within-batch offset for clip G (batch position i):
                    #   = (abs_desired[G] - abs_desired[batch_start]) - transition_duration
                    # This ensures: scene G fully visible at abs_desired[G] in the BATCH output.
                    # When batches are concatenated, the batch starts at the right position
                    # only if clip durations sum correctly, which requires padding.
                    filter_parts = []
                    prev_label = "[0]"

                    for i in range(1, len(batch_clips)):
                        xfade_type = get_xfade_type(video_transition)
                        global_idx = batch_start_index + i
                        batch_base_global = batch_start_index  # global index of batch's first clip

                        if global_idx in xfade_srt_offsets and batch_base_global in xfade_srt_offsets:
                            abs_time_G = xfade_srt_offsets[global_idx]      # fully visible abs time
                            abs_time_B = xfade_srt_offsets[batch_base_global]  # batch start abs time
                            within_batch_visible = abs_time_G - abs_time_B  # relative to batch start
                            current_offset = max(0.01, within_batch_visible - transition_duration)
                        else:
                            # Fallback: cumulative measured durations
                            current_offset = max(0.01, sum(batch_durations[:i]) - i * transition_duration)

                        out_label = f"[v{i}]" if i < len(batch_clips) - 1 else "[vout]"
                        filter_parts.append(
                            f"{prev_label}[{i}]xfade=transition={xfade_type}:duration={transition_duration}:offset={current_offset:.3f}{out_label}"
                        )
                        prev_label = out_label

                    filter_complex = ";".join(filter_parts) + ";[vout]format=yuv420p[vfinal]"

                    batch_output = Path(temp_dir_path) / f"batch_{batch_idx:03d}.mp4"

                    # Try NVENC first (3x faster), fall back to CPU libx264 on failure
                    # Clips are at render_size (1080p) - no 4K issues
                    def _run_xfade(enc, quality_args):
                        cmd = ["ffmpeg", "-y"] + inputs + [
                            "-filter_complex", filter_complex,
                            "-map", "[vfinal]",
                            "-c:v", enc, *quality_args,
                            "-pix_fmt", "yuv420p",
                            "-r", str(output_fps),
                            str(batch_output)
                        ]
                        return subprocess.run(cmd, capture_output=True, text=True, timeout=1200, creationflags=SUBPROCESS_FLAGS)

                    try:
                        if use_gpu:
                            result = _run_xfade(gpu_encoder, ["-preset", "p5", "-rc", "vbr", "-cq", "20", "-b:v", "15M"])
                            if result.returncode != 0:
                                plog(f"    xfade NVENC rc={result.returncode}, falling back to CPU...", "WARN")
                                if batch_output.exists(): batch_output.unlink()
                                result = _run_xfade("libx264", ["-preset", "fast", "-crf", "18", "-profile:v", "high"])
                        else:
                            result = _run_xfade("libx264", ["-preset", "fast", "-crf", "18", "-profile:v", "high"])

                        if result.returncode == 0 and _clip_is_valid(batch_output):
                            return batch_output
                        plog(f"    xfade batch returncode={result.returncode}: {result.stderr[-400:]}", "WARN")
                        try:
                            if batch_output.exists() and not _clip_is_valid(batch_output):
                                batch_output.unlink()
                        except:
                            pass
                        return None
                    except subprocess.TimeoutExpired:
                        plog(f"    xfade batch timed out (1200s)", "WARN")
                        return None

                # Build xfade_srt_offsets: maps clip_index -> xfade offset (in seconds).
                # Rule: scene N is FULLY VISIBLE at srt_start[N] in the video.
                # With xfade, transition STARTS at T seconds before full visibility:
                #   xfade_offset[N] = desired_fully_visible_time[N] - T
                #
                # For batch processing: offset is RELATIVE to batch start clip time.
                # batch_abs_times[i] = absolute video time of clip i's desired appearance.
                # within_batch_offset[i] = (batch_abs_times[i] - batch_abs_times[B]) - T
                #
                # Disclaimer (clip 0) adds D=1s before scene content starts.
                # Scene 1 (clip 1) is desired fully-visible at absolute t=D.
                # Scene 2 (clip 2) is desired fully-visible at absolute t=D + srt_start[2].
                # Scene N (clip N) is desired fully-visible at absolute t=D + srt_start[N].

                D_abs = DISCLAIMER_DURATION if any(it.get('is_disclaimer') for it in media_items) else 0.0

                # Build absolute desired appearance times keyed by CLIP POSITION in clip_paths.
                # clip_media_indices[clip_pos] = original media_items index, handles skipped clips.
                abs_desired = {}  # clip_position -> absolute video time when scene fully visible
                for clip_pos, media_idx in enumerate(clip_media_indices):
                    item = media_items[media_idx]
                    if item.get('is_disclaimer'):
                        abs_desired[clip_pos] = 0.0
                    else:
                        srt_s = item.get('start', 0.0)
                        if srt_s < 0:
                            srt_s = 0.0
                        abs_desired[clip_pos] = D_abs + srt_s

                # xfade_srt_offsets uses clip_position keys (same as abs_desired)
                xfade_srt_offsets = abs_desired

                plog(f"  xfade offsets (srt_start-based): {len(xfade_srt_offsets)} clips")
                if len(xfade_srt_offsets) >= 3:
                    sample = [(i, f'{v:.2f}s') for i, v in sorted(abs_desired.items())[:3]]
                    plog(f"    Abs desired: {sample}")

                # Process in batches to avoid Windows command line length limit
                BATCH_SIZE = int(os.environ.get("VE3_BATCH_SIZE", 15))
                xfade_success = False

                if len(clip_paths) <= BATCH_SIZE:
                    # Small video - process all at once
                    plog(f"  Using xfade transitions ({video_transition})...")
                    try:
                        batch_output = xfade_batch(clip_paths, 0, temp_dir, batch_start_index=0)
                        if batch_output:
                            shutil.move(str(batch_output), str(temp_video))
                            xfade_success = True
                    except Exception as e:
                        plog(f"  xfade exception: {e}", "WARN")
                else:
                    # Large video - process in batches
                    plog(f"  Processing {len(clip_paths)} clips in batches of {BATCH_SIZE}...")
                    batch_outputs = []
                    for batch_idx in range(0, len(clip_paths), BATCH_SIZE):
                        batch_num = batch_idx // BATCH_SIZE + 1
                        batch_clips = clip_paths[batch_idx:batch_idx + BATCH_SIZE]
                        plog(f"    Batch {batch_num}: clips {batch_idx + 1}-{batch_idx + len(batch_clips)}")
                        batch_output = None
                        for attempt in range(2):  # Try up to 2 times
                            try:
                                batch_output = xfade_batch(batch_clips, batch_idx // BATCH_SIZE, temp_dir, batch_start_index=batch_idx)
                            except Exception as e:
                                plog(f"    Batch {batch_num} attempt {attempt+1} exception: {e}", "WARN")
                                batch_output = None
                            if batch_output:
                                break
                            if attempt == 0 and not batch_output:
                                plog(f"    Batch {batch_num} attempt 1 failed, retrying...", "WARN")
                                time.sleep(3)
                        if batch_output:
                            batch_outputs.append(batch_output)
                        else:
                            plog(f"    Batch {batch_num} failed after 2 attempts", "WARN")
                            break

                    if len(batch_outputs) == (len(clip_paths) + BATCH_SIZE - 1) // BATCH_SIZE:
                        # All batches successful - concat them
                        plog(f"  Combining {len(batch_outputs)} batches...")
                        batch_list = Path(temp_dir) / "batches.txt"
                        with open(batch_list, 'w', encoding='utf-8') as f:
                            for bp in batch_outputs:
                                f.write(f"file '{str(bp).replace(chr(92), '/')}'\n")
                        cmd_final = ["ffmpeg", "-y", "-f", "concat", "-safe", "0", "-i", str(batch_list), "-c", "copy", str(temp_video)]
                        result = subprocess.run(cmd_final, capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS)
                        if result.returncode == 0:
                            # Verify xfade output duration
                            probe_cmd = ["ffprobe", "-v", "error", "-show_entries", "format=duration",
                                         "-of", "default=noprint_wrappers=1:nokey=1", str(temp_video)]
                            probe_r = subprocess.run(probe_cmd, capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS)
                            xfade_dur = float(probe_r.stdout.strip()) if probe_r.stdout.strip() else 0
                            plog(f"  xfade duration: {xfade_dur:.1f}s (expected ~{total_duration:.1f}s)")
                            if xfade_dur >= total_duration * 0.5:
                                xfade_success = True
                            else:
                                plog(f"  xfade output too short, will use fallback", "WARN")

                if not xfade_success:
                    plog(f"  xfade failed, falling back to CPU re-encode concat", "WARN")
                    list_file = Path(temp_dir) / "clips.txt"
                    with open(list_file, 'w', encoding='utf-8') as f:
                        for cp in clip_paths:
                            f.write(f"file '{str(cp).replace(chr(92), '/')}'\n")
                    # Use libx264 re-encode (not -c copy) to normalize clips and ensure correct duration
                    rw, rh = render_size
                    cmd_concat = ["ffmpeg", "-y", "-f", "concat", "-safe", "0", "-i", str(list_file),
                                  "-vf", f"scale={rw}:{rh}:flags=lanczos,format=yuv420p",
                                  "-c:v", "libx264", "-preset", "fast", "-crf", "18",
                                  "-pix_fmt", "yuv420p", "-an", str(temp_video)]
                    result = subprocess.run(cmd_concat, capture_output=True, text=True, timeout=7200, creationflags=SUBPROCESS_FLAGS)
                    if result.returncode != 0:
                        return False, None, f"Concat error: {result.stderr[-200:]}"
                    # Verify concat duration
                    probe_cmd = ["ffprobe", "-v", "error", "-show_entries", "format=duration",
                                 "-of", "default=noprint_wrappers=1:nokey=1", str(temp_video)]
                    probe_r = subprocess.run(probe_cmd, capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS)
                    concat_dur = float(probe_r.stdout.strip()) if probe_r.stdout.strip() else 0
                    expected_dur = total_duration
                    plog(f"  Concat duration: {concat_dur:.1f}s (expected ~{expected_dur:.1f}s)")
                    if concat_dur < expected_dur * 0.5:
                        return False, None, f"Concat too short: {concat_dur:.1f}s vs expected {expected_dur:.1f}s"
            else:
                # Simple concat (for fade_black or single clip)
                list_file = Path(temp_dir) / "clips.txt"
                with open(list_file, 'w', encoding='utf-8') as f:
                    for cp in clip_paths:
                        f.write(f"file '{str(cp).replace(chr(92), '/')}'\n")

                cmd_concat = ["ffmpeg", "-y", "-f", "concat", "-safe", "0", "-i", str(list_file), "-c", "copy", str(temp_video)]
                result = subprocess.run(cmd_concat, capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS)
                if result.returncode != 0:
                    return False, None, f"Concat error: {result.stderr[-200:]}"

            _t['concat'] = time.time()

            # Fix xfade duration mismatch: xfade transitions make video shorter than voice
            # Efficiently pad: extract last PNG frame → encode short freeze clip → concat (zero-copy)
            probe_vid = subprocess.run(
                ["ffprobe", "-v", "error", "-show_entries", "format=duration",
                 "-of", "default=noprint_wrappers=1:nokey=1", str(temp_video)],
                capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS
            )
            actual_vid_dur = float(probe_vid.stdout.strip()) if probe_vid.stdout.strip() else total_duration
            pad_needed = total_duration - actual_vid_dur
            if pad_needed > 10.0:
                plog(f"  Extending video: {actual_vid_dur:.1f}s -> {total_duration:.1f}s (freeze {pad_needed:.1f}s)")
                fw, fh = output_size  # temp_video is at render_size (1080p), upscale later in subtitle burn
                last_frame_png = Path(temp_dir) / "last_frame.png"
                freeze_clip = Path(temp_dir) / "freeze.mp4"
                # Extract last frame as PNG
                ext_r = subprocess.run(
                    ["ffmpeg", "-y", "-sseof", "-0.5", "-i", str(temp_video),
                     "-vframes", "1", str(last_frame_png)],
                    capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS
                )
                if ext_r.returncode == 0 and last_frame_png.exists():
                    # Must match xfade_batch output codec (always libx264) for stream copy concat
                    freeze_enc = ["-c:v", "libx264", "-preset", "fast", "-crf", "18", "-profile:v", "high"]
                    frz_r = subprocess.run(
                        ["ffmpeg", "-y", "-loop", "1", "-framerate", str(output_fps),
                         "-i", str(last_frame_png), "-t", f"{pad_needed:.2f}",
                         "-vf", f"scale={fw}:{fh}:flags=lanczos",
                         "-pix_fmt", "yuv420p", "-r", str(output_fps)] + freeze_enc + [str(freeze_clip)],
                        capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS
                    )
                    if frz_r.returncode == 0 and freeze_clip.exists():
                        # Concat original + freeze (zero-copy, fast)
                        temp_video_extended = Path(temp_dir) / "extended.mp4"
                        list_ext = Path(temp_dir) / "extend_list.txt"
                        with open(list_ext, 'w') as lf:
                            lf.write(f"file '{str(temp_video).replace(chr(92), '/')}'\n")
                            lf.write(f"file '{str(freeze_clip).replace(chr(92), '/')}'\n")
                        ext2_r = subprocess.run(
                            ["ffmpeg", "-y", "-f", "concat", "-safe", "0",
                             "-i", str(list_ext), "-c", "copy", str(temp_video_extended)],
                            capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS
                        )
                        if ext2_r.returncode == 0:
                            temp_video = temp_video_extended
                            plog(f"  Extended OK")
                        else:
                            plog(f"  Concat extend failed: {ext2_r.stderr[-100:]}", "WARN")
                    else:
                        plog(f"  Freeze encode failed: {frz_r.stderr[-100:]}", "WARN")
                else:
                    plog(f"  Last frame extract failed: {ext_r.stderr[-100:]}", "WARN")

            # Add audio (voice + optional background music)
            temp_with_audio = Path(temp_dir) / "with_audio.mp4"
            update_progress(code=code, step="Adding audio", percent=85)

            # ── Background music logic ────────────────────────────────────────
            # If project_dir/music/ exists with .mp3 files AND Excel has a 'music'
            # sheet with start_time data, mix each track at the correct timestamp.
            # Each track is trimmed so it ends exactly when the next track starts.
            # Keep music clearly in the background: lower base level + duck under voice.
            MUSIC_VOLUME = 0.22   # 22% base bed: clearer background music, still ducked under voice
            MUSIC_DUCK_THRESHOLD = 0.035
            MUSIC_DUCK_RATIO = 10
            music_dir    = project_dir / "music"
            music_segments = []  # list of (start_sec, allowed_dur_sec, mp3_path)

            if music_dir.exists() and any(music_dir.glob("*.mp3")):
                # ── Read timing from Excel 'music' sheet ──────────────────────
                try:
                    wb_m = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
                    if "music" in wb_m.sheetnames:
                        ws_m   = wb_m["music"]
                        m_hdrs = [str(c.value).lower().strip() if c.value else ''
                                  for c in ws_m[1]]
                        def _col(name):
                            return m_hdrs.index(name) if name in m_hdrs else None
                        col_id     = _col('music_id')
                        col_start  = _col('start_time')
                        col_status = _col('status')

                        raw_segs = []
                        for row in ws_m.iter_rows(min_row=2, values_only=True):
                            if col_id is None or row[col_id] is None:
                                continue
                            status = (str(row[col_status]).lower().strip()
                                      if col_status is not None and row[col_status]
                                      else 'pending')
                            if status == 'skip':
                                continue
                            start_s = (parse_timestamp(str(row[col_start]))
                                       if col_start is not None and row[col_start]
                                       else 0.0)
                            try:
                                mid = str(int(float(str(row[col_id]))))
                            except Exception:
                                continue
                            mp3 = music_dir / f"{mid}.mp3"
                            if not mp3.exists():
                                alts = list(music_dir.glob(f"{mid}.*"))
                                mp3  = alts[0] if alts else None
                            raw_segs.append({'start': start_s, 'path': mp3})

                        wb_m.close()
                        raw_segs.sort(key=lambda x: x['start'])

                        for i, seg in enumerate(raw_segs):
                            if not seg['path'] or not seg['path'].exists():
                                continue
                            # Trim track so it ends when next track begins
                            if i < len(raw_segs) - 1:
                                allowed = raw_segs[i + 1]['start'] - seg['start']
                            else:
                                allowed = total_duration - seg['start']
                            allowed = max(allowed, 5.0)
                            music_segments.append((seg['start'], allowed, seg['path']))
                        plog(f"  Music: {len(music_segments)} tracks (from Excel sheet)")
                except Exception as _em:
                    plog(f"  Music sheet read error (skipping music): {_em}", "WARN")
                    music_segments = []

                # ── Fallback: auto-detect from sorted filenames ───────────────
                if not music_segments:
                    mp3s = sorted(music_dir.glob("*.mp3"),
                                  key=lambda f: (int(f.stem) if f.stem.isdigit() else 999))
                    n_mp3   = len(mp3s)
                    seg_dur = total_duration / n_mp3 if n_mp3 > 0 else total_duration
                    for i, mp3 in enumerate(mp3s):
                        music_segments.append((i * seg_dur, seg_dur, mp3))
                    plog(f"  Music: {len(music_segments)} files (auto-spaced, no sheet)")

            # ── Mix music tracks → single music_mix.mp3 ───────────────────────
            if music_segments:
                try:
                    music_mix_path = Path(temp_dir) / "music_mix.mp3"
                    m_inputs, m_filters, m_labels = [], [], []
                    for idx, (start_sec, allowed_dur, mp3_path) in enumerate(music_segments):
                        m_inputs.extend(["-i", str(mp3_path)])
                        delay_ms = int(start_sec * 1000)
                        lbl = f"[m{idx}]"
                        m_filters.append(
                            f"[{idx}:a]atrim=0:{allowed_dur:.3f},"
                            f"asetpts=PTS-STARTPTS,"
                            f"adelay={delay_ms}|{delay_ms}{lbl}"
                        )
                        m_labels.append(lbl)

                    n_m = len(m_labels)
                    m_filters.append(
                        # normalize=0: sum tracks without dividing (sequential = no clipping)
                        f"{''.join(m_labels)}amix=inputs={n_m}:duration=longest:normalize=0[mout]"
                    )
                    cmd_music = (["ffmpeg", "-y"] + m_inputs + [
                        "-filter_complex", ";".join(m_filters),
                        "-map", "[mout]", "-b:a", "192k", str(music_mix_path)
                    ])
                    r_m = subprocess.run(cmd_music, capture_output=True, text=True,
                                         timeout=300, creationflags=SUBPROCESS_FLAGS)

                    if r_m.returncode == 0 and music_mix_path.exists() \
                            and music_mix_path.stat().st_size > 1000:
                        # ── Mix voice + background music with speech-priority ducking ──
                        temp_mixed = Path(temp_dir) / "voice_with_music.mp3"
                        cmd_amix = [
                            "ffmpeg", "-y",
                            "-i", str(voice_path),
                            "-i", str(music_mix_path),
                            "-filter_complex",
                            # Lower music first, then duck it further whenever voice is active.
                            f"[1:a]volume={MUSIC_VOLUME}[bgbase];"
                            f"[bgbase][0:a]sidechaincompress=threshold={MUSIC_DUCK_THRESHOLD}:"
                            f"ratio={MUSIC_DUCK_RATIO}:attack=20:release=350:makeup=1"
                            f"[bgduck];"
                            f"[0:a][bgduck]amix=inputs=2:duration=first:"
                            f"weights=1.0 1.0:normalize=0[aout]",
                            "-map", "[aout]", "-b:a", "256k", str(temp_mixed)
                        ]
                        r_amix = subprocess.run(cmd_amix, capture_output=True, text=True,
                                                 timeout=600, creationflags=SUBPROCESS_FLAGS)
                        if r_amix.returncode == 0 and temp_mixed.exists():
                            voice_path = temp_mixed  # use mixed audio downstream
                            plog(f"  Voice + music mixed OK "
                                 f"(voice priority, music base {int(MUSIC_VOLUME*100)}%, ducked under speech)")
                        else:
                            plog(f"  amix voice+music failed: {r_amix.stderr[-150:]}", "WARN")
                    else:
                        plog(f"  Music track build failed "
                             f"(rc={r_m.returncode}): {r_m.stderr[-200:]}", "WARN")
                except Exception as _emx:
                    plog(f"  Music mix error (continuing without music): {_emx}", "WARN")
            # ─────────────────────────────────────────────────────────────────

            # Final audio master after optional music mix. This is intentionally after
            # ducking/amix, because mixing changes integrated loudness and true peak.
            final_audio = normalize_voice(voice_path, Path(temp_dir), label="final_audio")
            if final_audio != voice_path:
                voice_path = final_audio
                plog("  Final audio mastered for YouTube loudness")
            else:
                plog("  Final audio master fallback: using current audio", "WARN")

            plog("  Adding audio to video...")
            # When disclaimer is prepended (D seconds), delay voice to stay in sync with video scenes.
            # Without delay: voice starts at t=0 but scene 1 video appears at t=D → audio leads by D.
            if disclaimer_img:
                delay_ms = int(DISCLAIMER_DURATION * 1000)
                cmd2 = ["ffmpeg", "-y", "-i", str(temp_video), "-i", str(voice_path),
                        "-filter_complex",
                        f"[1:a]adelay={delay_ms}|{delay_ms}[adelayed]",
                        "-map", "0:v", "-map", "[adelayed]",
                        "-c:v", "copy", "-c:a", "aac", "-b:a", "256k", "-shortest", str(temp_with_audio)]
                plog(f"  Audio delayed {DISCLAIMER_DURATION}s to sync with disclaimer")
            else:
                cmd2 = ["ffmpeg", "-y", "-i", str(temp_video), "-i", str(voice_path),
                       "-c:v", "copy", "-c:a", "aac", "-b:a", "256k", "-shortest", str(temp_with_audio)]
            result = subprocess.run(cmd2, capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS)
            if result.returncode != 0:
                return False, None, f"Audio merge error: {result.stderr[-200:]}"
            _t['audio'] = time.time()

            # Burn subtitles
            temp_with_subs = Path(temp_dir) / "with_subs.mp4"
            if srt_path and srt_path.exists():
                update_progress(code=code, step="Burning subtitles", percent=90)

                # Default behavior is voice sync: only shift by disclaimer.
                # Optional "video" mode keeps xfade compensation behavior.
                burn_srt = srt_path
                subtitle_sync_mode = str(channel_template.get("subtitle_sync", "voice")).lower().strip()
                if subtitle_sync_mode == "video" and use_xfade:
                    shifted_srt = Path(temp_dir) / f"{srt_path.stem}_xfade_adjusted.srt"
                    burn_srt = shift_srt_with_xfade_compensation(
                        srt_path, media_items, transition_duration,
                        DISCLAIMER_DURATION if disclaimer_img else 0.0,
                        shifted_srt
                    )
                    plog(f"  SRT adjusted for xfade transitions (T={transition_duration}s)")
                elif disclaimer_img and DISCLAIMER_DURATION > 0:
                    shifted_srt = Path(temp_dir) / f"{srt_path.stem}_shifted.srt"
                    burn_srt = shift_srt(srt_path, DISCLAIMER_DURATION, shifted_srt)
                    plog(f"  SRT shifted +{DISCLAIMER_DURATION}s to match disclaimer")
                elif subtitle_sync_mode not in ("voice", "video"):
                    plog(f"  Unknown subtitle_sync='{subtitle_sync_mode}', fallback to voice sync", "WARN")

                # NVENC for subtitle burn: libass renders on CPU, encoding offloaded to GPU
                # Test proved: NVENC = ~17 min vs CPU = ~49 min (3x faster)
                use_gpu_sub = use_gpu
                sub_method = "NVENC" if use_gpu_sub else "CPU"
                if _parallel_count >= 2:
                    plog(f"  Burning subtitles [{sub_method}] (slot {_parallel_slot})...")
                else:
                    plog(f"  Burning subtitles [{sub_method}]...")

                # Use local fonts from fonts/ folder with channel template
                fonts_dir = str(TOOL_DIR / "fonts").replace('\\', '/').replace(':', '\\:')
                template = get_subtitle_template(code)
                requested_font = template.get("font", "Auto")
                subtitle_text = read_text_best_effort(burn_srt)
                resolved_font = choose_subtitle_font(requested_font, subtitle_text)
                detected_lang = detect_subtitle_language(subtitle_text)
                render_profile = get_subtitle_render_profile(resolved_font, detected_lang)
                processed_srt = Path(temp_dir) / f"{burn_srt.stem}_wrapped.srt"
                burn_srt = process_srt_for_video(burn_srt, processed_srt, max_chars=render_profile["max_chars"])
                utf8_srt = Path(temp_dir) / f"{burn_srt.stem}_utf8.srt"
                burn_srt = normalize_srt_utf8(burn_srt, utf8_srt)
                if not srt_has_cues(burn_srt):
                    return False, None, f"Subtitle error: SRT has no cues or is empty ({burn_srt})"

                srt_escaped = str(burn_srt).replace('\\', '/').replace(':', '\\:')
                if resolved_font != requested_font:
                    plog(f"  Subtitle font selected [{detected_lang}]: {requested_font} -> {resolved_font}")
                elif requested_font == "Auto":
                    plog(f"  Subtitle font selected [{detected_lang}]: {resolved_font}")
                font_size, outline_size, margin_v = subtitle_style_numbers(template, resolved_font, detected_lang)
                if font_size != _to_int(template.get("size", 28), 28) or render_profile["max_chars"] != 45:
                    plog(f"  Subtitle fit: size={font_size}, max_chars={render_profile['max_chars']}, margin_v={margin_v}")
                subtitle_style = (
                    f"FontName={escape_ass_style_value(resolved_font)},FontSize={font_size},"
                    f"PrimaryColour={escape_ass_style_value(template['color'])},OutlineColour={escape_ass_style_value(template['outline'])},"
                    f"BorderStyle=1,Outline={outline_size},Shadow=1,WrapStyle=2,"
                    f"MarginV={margin_v},Alignment={template['alignment']}"
                )
                vf_filter = f"subtitles='{srt_escaped}':fontsdir='{fonts_dir}':charenc=UTF-8:force_style='{subtitle_style}'"

                # Upscale 1080p → 4K combined with subtitle burn (single pass, efficient)
                if needs_upscale:
                    fw, fh = final_output_size
                    vf_filter = f"scale={fw}:{fh}:flags=lanczos," + vf_filter

                if use_gpu_sub:
                    cmd3 = ["ffmpeg", "-y", "-i", str(temp_with_audio),
                            "-vf", vf_filter,
                            "-c:v", gpu_encoder, "-preset", "p5", "-rc", "vbr", "-cq", "20", "-b:v", "15M",
                            "-c:a", "copy", str(temp_with_subs)]
                else:
                    cmd3 = ["ffmpeg", "-y", "-i", str(temp_with_audio), "-vf", vf_filter, "-c:a", "copy", str(temp_with_subs)]
                result = subprocess.run(cmd3, capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS)
                if result.returncode != 0:
                    plog(f"  Subtitle burn [{sub_method}] failed (rc={result.returncode}): {result.stderr[-300:]}", "WARN")
                    if use_gpu_sub:
                        # Fallback to CPU encode
                        plog("  Retrying subtitle burn with CPU...", "WARN")
                        cmd3_cpu = ["ffmpeg", "-y", "-i", str(temp_with_audio), "-vf", vf_filter, "-c:a", "copy", str(temp_with_subs)]
                        result2 = subprocess.run(cmd3_cpu, capture_output=True, text=True, creationflags=SUBPROCESS_FLAGS)
                        if result2.returncode != 0:
                            plog(f"  CPU subtitle burn also failed: {result2.stderr[-200:]}", "ERROR")
                            return False, None, f"Subtitle burn error: {result2.stderr[-500:]}"
                    else:
                        return False, None, f"Subtitle burn error: {result.stderr[-500:]}"
            else:
                if needs_upscale:
                    fw, fh = final_output_size
                    enc_args = ["-c:v", gpu_encoder, "-preset", "p5", "-rc", "vbr", "-cq", "20", "-b:v", "15M"] if use_gpu else ["-c:v", "libx264", "-preset", "fast", "-crf", "17"]
                    subprocess.run(["ffmpeg", "-y", "-i", str(temp_with_audio),
                                    "-vf", f"scale={fw}:{fh}:flags=lanczos"] + enc_args + ["-c:a", "copy", str(temp_with_subs)],
                                   capture_output=True, creationflags=SUBPROCESS_FLAGS)
                else:
                    shutil.copy(temp_with_audio, temp_with_subs)
            _t['subs'] = time.time()

            # Overlay NV image (character card) if enabled in template
            nv_enabled = channel_template.get("nv_overlay_enabled", True)
            nv_position = channel_template.get("nv_overlay_position", "left")
            nv_v_position = channel_template.get("nv_overlay_v_position", "middle")
            nv_scale = channel_template.get("nv_overlay_scale", 0.50)
            nv_crop_ratio = channel_template.get("nv_crop_ratio", 0.5)

            nv_path = find_nv_image(code, project_dir) if nv_enabled else None
            if nv_path:
                update_progress(code=code, step="Adding NV overlay", percent=95)
                plog(f"  Adding NV overlay: {nv_path.name} ({nv_position}-{nv_v_position}, {int(nv_scale*100)}%, crop={nv_crop_ratio})")
                if overlay_nv_on_video(temp_with_subs, nv_path, output_path,
                                       position=nv_position, v_position=nv_v_position,
                                       scale=nv_scale, margin=20, crop_ratio=nv_crop_ratio,
                                       callback=callback):
                    plog("  NV overlay applied successfully")
                else:
                    # Fallback: use video without NV overlay
                    shutil.copy(temp_with_subs, output_path)
            else:
                # No NV image or disabled, just copy the video with subtitles
                shutil.copy(temp_with_subs, output_path)

            _t['nv'] = time.time()

            update_progress(code=code, step="Done", percent=100, status="completed")
            plog(f"  Video done: {output_path.name}", "OK")

            # Lưu nhật ký thời gian
            def _dur(k1, k2):
                return round(_t.get(k2, _t['nv']) - _t.get(k1, _t['start']), 1)
            steps = {
                'clip_creation':   _dur('start',   'clips'),
                'reencode':        _dur('clips',   'reencode'),
                'concat_xfade':    _dur('reencode','concat'),
                'audio':           _dur('concat',  'audio'),
                'subtitle_burn':   _dur('audio',   'subs'),
                'nv_overlay':      _dur('subs',    'nv'),
                'total':           _dur('start',   'nv'),
            }
            save_timing_log({
                'timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
                'code': code,
                'num_clips': _num_clips,
                'output_resolution': output_resolution,
                'use_gpu': use_gpu,
                'clip_workers': clip_workers,
                'transition': 'xfade' if use_xfade else 'fade',
                'steps_s': steps,
            })
            total_min = steps['total'] / 60
            def _m(k): return f"{steps[k]/60:.0f}p" if steps[k] >= 60 else f"{steps[k]:.0f}s"
            plog(f"  Thoi gian: {total_min:.0f} phut | Clip {_m('clip_creation')} | Reencode {_m('reencode')} | Concat {_m('concat_xfade')} | Sub {_m('subtitle_burn')} | NV {_m('nv_overlay')}", "OK")

            return True, output_path, None

        finally:
            gc.collect()
            time.sleep(1)
            for attempt in range(5):
                try:
                    shutil.rmtree(temp_dir, ignore_errors=False)
                    break
                except PermissionError:
                    gc.collect()
                    time.sleep(1)
                    if attempt == 4:
                        shutil.rmtree(temp_dir, ignore_errors=True)

    except Exception as e:
        plog(f"  Video compose error: {e}", "ERROR")
        import traceback
        traceback.print_exc()
        return False, None, str(e)


# ============================================================================
# COPY TO DONE
# ============================================================================

def find_thumbnail(code: str) -> Optional[Path]:
    if not THUMB_DIR.exists():
        return None
    for ext in [".png", ".jpg", ".jpeg", ".webp"]:
        thumb = THUMB_DIR / f"{code}{ext}"
        if thumb.exists():
            return thumb
    return None


def find_nv_image(code: str, project_dir: Path) -> Optional[Path]:
    """Find NV image (character card) for overlay."""
    # Check in VISUAL project folder first
    nv_in_project = project_dir / f"{code}_nv.png"
    if nv_in_project.exists():
        return nv_in_project

    # Check in thumb/nv folder
    nv_in_thumb = TOOL_DIR / "thumb" / "nv" / f"{code}.png"
    if nv_in_thumb.exists():
        return nv_in_thumb

    return None


def overlay_nv_on_video(video_path: Path, nv_path: Path, output_path: Path,
                        position: str = "left", v_position: str = "middle",
                        scale: float = 0.50, margin: int = 20,
                        crop_ratio: float = 0.5, callback=None) -> bool:
    """
    Overlay NV image on video.

    Args:
        video_path: Input video
        nv_path: NV image (character card with name badge)
        output_path: Output video with overlay
        position: "left" or "right" (horizontal)
        v_position: "top", "middle", or "bottom" (vertical)
        scale: Scale factor for NV image (0.50 = 50% of video height)
        margin: Margin from edge in pixels
        crop_ratio: Crop right portion of NV image (0.5 = right half, 1.0 = full image)
    """
    def plog(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            log(msg, level)

    try:
        # Get video dimensions
        probe_cmd = [
            "ffprobe", "-v", "error",
            "-select_streams", "v:0",
            "-show_entries", "stream=width,height",
            "-of", "csv=p=0:s=x",
            str(video_path)
        ]
        result = subprocess.run(probe_cmd, capture_output=True, text=True,
                               creationflags=SUBPROCESS_FLAGS)
        if result.returncode != 0:
            return False

        dimensions = result.stdout.strip()
        if 'x' not in dimensions:
            return False

        vid_w, vid_h = map(int, dimensions.split('x'))

        # Calculate NV overlay size (scale relative to video height)
        nv_height = int(vid_h * scale)

        # Build filter: scale NV, then overlay
        nv_escaped = str(nv_path).replace('\\', '/').replace(':', '\\:')

        # Horizontal position
        if position == "left":
            x_pos = margin
        else:
            x_pos = f"W-w-{margin}"

        # Vertical position
        if v_position == "top":
            y_pos = margin
        elif v_position == "middle":
            y_pos = "(H-h)/2"
        else:  # bottom
            y_pos = f"H-h-{margin}"

        # Filter: crop right portion (if needed), scale NV to height, then overlay
        if crop_ratio < 1.0:
            # Crop to right portion: crop=width:height:x:y
            crop_width = f"iw*{crop_ratio}"
            crop_x = f"iw*{1.0 - crop_ratio}"
            crop_filter = f"crop={crop_width}:ih:{crop_x}:0,"
        else:
            crop_filter = ""

        filter_complex = (
            f"[1:v]{crop_filter}scale=-1:{nv_height}[nv];"
            f"[0:v][nv]overlay={x_pos}:{y_pos}"
        )

        # Detect GPU encoder
        use_nvenc = False
        try:
            gpu_check = subprocess.run(["ffmpeg", "-encoders"], capture_output=True, text=True, timeout=5, creationflags=SUBPROCESS_FLAGS)
            use_nvenc = "h264_nvenc" in gpu_check.stdout
        except:
            pass

        if use_nvenc:
            cmd = [
                "ffmpeg", "-y",
                "-i", str(video_path),
                "-i", str(nv_path),
                "-filter_complex", filter_complex,
                "-c:a", "copy",
                "-c:v", "h264_nvenc", "-preset", "p4", "-rc", "vbr", "-cq", "20",
                str(output_path)
            ]
        else:
            cmd = [
                "ffmpeg", "-y",
                "-i", str(video_path),
                "-i", str(nv_path),
                "-filter_complex", filter_complex,
                "-c:a", "copy",
                "-c:v", "libx264", "-preset", "fast", "-crf", "20",
                str(output_path)
            ]

        result = subprocess.run(cmd, capture_output=True, text=True,
                               timeout=300, creationflags=SUBPROCESS_FLAGS)

        if result.returncode == 0:
            plog(f"  NV overlay added: {position}-{v_position}")
            return True
        else:
            plog(f"  NV overlay failed: {result.stderr[-100:]}", "WARN")
            return False

    except Exception as e:
        plog(f"  NV overlay error: {e}", "WARN")
        return False


def select_best_thumbnail_for_done(project_info: Dict, callback=None) -> bool:
    """Create VISUAL/<code>/<code>.jpg from VISUAL/<code>/thumb when available."""
    code = project_info["code"]
    project_dir = project_info["path"]

    def plog(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            log(f"[{code}] {msg}", level)

    visual_thumb_dir = project_dir / "thumb"
    if not visual_thumb_dir.exists() or not visual_thumb_dir.is_dir():
        return False

    selector_script = TOOL_DIR / "select_best_thumb.py"
    if not selector_script.exists():
        plog("Thumbnail selector not found: select_best_thumb.py", "WARN")
        return False

    try:
        result = subprocess.run(
            [
                sys.executable,
                str(selector_script),
                "--code", code,
                "--project-dir", str(project_dir),
                "--output", str(project_dir / f"{code}.jpg"),
            ],
            capture_output=True,
            text=True,
            timeout=180,
            encoding="utf-8",
            errors="replace",
            creationflags=SUBPROCESS_FLAGS if sys.platform == "win32" else 0,
        )

        output = (result.stdout or result.stderr or "").strip()
        if result.returncode == 0:
            if output:
                plog(output.splitlines()[-1])
            return True

        if result.returncode == 2:
            plog("No images in VISUAL thumb folder to select", "WARN")
        else:
            plog(f"Thumbnail selector failed: {output[-300:] if output else 'Unknown error'}", "WARN")
        return False

    except subprocess.TimeoutExpired:
        plog("Thumbnail selector timed out", "WARN")
        return False
    except Exception as e:
        plog(f"Thumbnail selector error: {e}", "WARN")
        return False


def find_fallback_thumbnail_source(project_info: Dict) -> Optional[Path]:
    """Find a usable project image when the dedicated thumbnail pipeline has no output."""
    project_dir = project_info["path"]
    code = project_info["code"]

    def _natural_key(name: str):
        return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", name)]

    candidates = [
        project_dir / f"{code}.jpg",
        project_dir / f"{code}.png",
        project_dir / "thumbnail" / "thumb_003.png",
        project_dir / "thumb" / "thumb_003.png",
    ]

    for base in [project_dir / "thumbnail", project_dir / "thumb", project_dir / "img", project_dir / "img_backup", project_dir]:
        if not base.exists() or not base.is_dir():
            continue
        for ext in ("*.jpg", "*.jpeg", "*.png", "*.webp"):
            candidates.extend(sorted(base.glob(ext), key=lambda p: _natural_key(p.name)))

    for src in candidates:
        if src.exists() and src.is_file():
            stem = src.stem.lower()
            if stem.startswith(("nv", "loc")):
                continue
            return src

    return None


def create_fallback_thumbnail(project_info: Dict, dst_path: Path, callback=None) -> bool:
    """Create required DONE JPG from project media when no explicit thumbnail exists."""
    code = project_info["code"]
    project_dir = project_info["path"]

    def plog(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            log(f"[{code}] {msg}", level)

    src_image = find_fallback_thumbnail_source(project_info)
    if src_image:
        if copy_thumbnail_as_jpg(src_image, dst_path):
            plog(f"Created fallback thumbnail from {src_image.parent.name}/{src_image.name}: {dst_path.name}", "WARN")
            return True
        plog(f"Fallback thumbnail image could not be normalized: {src_image}", "WARN")

    video_path = project_dir / f"{code}.mp4"
    if not video_path.exists():
        return False

    try:
        dst_path.parent.mkdir(parents=True, exist_ok=True)
        frame_path = dst_path.parent / f"{code}_frame_tmp.jpg"
        cmd = [
            "ffmpeg", "-y",
            "-ss", "3",
            "-i", str(video_path),
            "-frames:v", "1",
            "-q:v", "3",
            str(frame_path),
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120, creationflags=SUBPROCESS_FLAGS)
        if result.returncode != 0 or not frame_path.exists():
            plog(f"Fallback thumbnail frame failed: {result.stderr[-200:] if result.stderr else 'Unknown error'}", "WARN")
            return False
        ok = copy_thumbnail_as_jpg(frame_path, dst_path)
        try:
            frame_path.unlink()
        except Exception:
            pass
        if ok:
            plog(f"Created fallback thumbnail from rendered video: {dst_path.name}", "WARN")
        return ok
    except Exception as e:
        plog(f"Fallback thumbnail error: {e}", "WARN")
        return False


def copy_thumbnail_as_jpg(src_path: Path, dst_path: Path, max_bytes: int = 2 * 1024 * 1024) -> bool:
    """Copy/convert a thumbnail source to JPG and keep it under max_bytes."""
    try:
        from PIL import Image, ImageEnhance, ImageOps

        img = Image.open(src_path)
        img = ImageOps.exif_transpose(img)
        if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
            bg = Image.new("RGB", img.size, "white")
            rgba = img.convert("RGBA")
            bg.paste(rgba, mask=rgba.split()[-1])
            img = bg
        else:
            img = img.convert("RGB")

        img.thumbnail((1920, 1080), Image.Resampling.LANCZOS)
        img = ImageOps.autocontrast(img, cutoff=0.4)
        img = ImageEnhance.Sharpness(img).enhance(1.08)

        dst_path.parent.mkdir(parents=True, exist_ok=True)
        work = img
        for _ in range(5):
            for quality in (95, 92, 90, 88, 85, 82, 78, 74, 70, 66, 62, 58):
                work.save(dst_path, "JPEG", quality=quality, optimize=True, progressive=True, subsampling=1)
                if dst_path.stat().st_size <= max_bytes:
                    return True
            next_size = (max(640, int(work.width * 0.88)), max(360, int(work.height * 0.88)))
            if next_size == work.size:
                break
            work = work.resize(next_size, Image.Resampling.LANCZOS)

        work.save(dst_path, "JPEG", quality=55, optimize=True, progressive=True, subsampling=2)
        return dst_path.exists() and dst_path.stat().st_size <= max_bytes
    except Exception:
        return False


def validate_done_folder(done_folder: Path, require_thumb_folder: bool = False) -> Tuple[bool, List[str]]:
    """Validate final DONE/<code> package has required deliverables."""
    missing = []

    if not any(done_folder.glob("*.mp4")):
        missing.append("mp4")
    if not any(done_folder.glob("*.srt")):
        missing.append("srt")

    jpg_files = list(done_folder.glob("*.jpg"))
    if not jpg_files:
        missing.append("jpg")
    elif not any(p.stat().st_size <= 2 * 1024 * 1024 for p in jpg_files):
        missing.append("jpg < 2MB")

    thumb_dir = done_folder / "thumb"
    if require_thumb_folder and (not thumb_dir.exists() or not thumb_dir.is_dir()):
        missing.append("thumb folder")

    return len(missing) == 0, missing


def run_seo_tho_for_done(project_info: Dict, done_folder: Path, callback=None) -> bool:
    """Rename DONE mp4/jpg/srt by clean video title and write basic metadata."""
    code = project_info["code"]
    project_dir = project_info["path"]

    def plog(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            log(f"[{code}] {msg}", level)

    seo_script = TOOL_DIR / "seo_tho.py"
    if not seo_script.exists():
        plog("SEO-tho script not found: seo_tho.py", "WARN")
        return False

    try:
        result = subprocess.run(
            [
                sys.executable,
                str(seo_script),
                "--code", code,
                "--done-dir", str(done_folder),
                "--project-dir", str(project_dir),
            ],
            capture_output=True,
            text=True,
            timeout=900,
            encoding="utf-8",
            errors="replace",
            creationflags=SUBPROCESS_FLAGS if sys.platform == "win32" else 0,
        )
        output = (result.stdout or result.stderr or "").strip()
        if result.returncode == 0:
            if output:
                plog(output.splitlines()[-1])
            return True

        plog(f"SEO-tho failed: {output[-400:] if output else 'Unknown error'}", "WARN")
        return False
    except subprocess.TimeoutExpired:
        plog("SEO-tho timed out", "WARN")
        return False
    except Exception as e:
        plog(f"SEO-tho error: {e}", "WARN")
        return False


def copy_to_done(project_info: Dict, video_path: Path, callback=None) -> Tuple[bool, Optional[str]]:
    code = project_info["code"]
    project_dir = project_info["path"]

    def plog(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            log(f"[{code}] {msg}", level)

    done_folder = DONE_DIR / code

    if done_folder.exists():
        plog("Removing old done folder...")
        shutil.rmtree(done_folder)

    done_folder.mkdir(parents=True, exist_ok=True)
    plog(f"Created: {done_folder}")

    # 1. Copy video
    dst_video = done_folder / video_path.name
    shutil.copy2(video_path, dst_video)
    plog(f"Copied video: {dst_video.name}")

    # 2. Copy SRT
    srt_path = project_info.get("srt_path")
    if srt_path and srt_path.exists():
        dst_srt = done_folder / f"{code}.srt"
        shutil.copy2(srt_path, dst_srt)
        plog(f"Copied SRT: {dst_srt.name}")

    # 3. Select/normalize thumbnail from VISUAL/<code>/thumb before copying.
    select_best_thumbnail_for_done(project_info, callback)

    # 4. Copy thumbnail (check multiple locations)
    thumb_copied = False
    dst_thumb = done_folder / f"{code}_thumb.jpg"

    # Check VISUAL folder first (generated by run_thumb.py)
    thumb_in_visual = project_dir / f"{code}.jpg"
    if thumb_in_visual.exists():
        if copy_thumbnail_as_jpg(thumb_in_visual, dst_thumb):
            plog(f"Copied thumbnail: {dst_thumb.name}")
            thumb_copied = True
        else:
            plog(f"Cannot normalize thumbnail: {thumb_in_visual.name}", "WARN")

    # Check thumb/thumbnails folder
    if not thumb_copied:
        thumb_in_tool = TOOL_DIR / "thumb" / "thumbnails" / f"{code}.jpg"
        if thumb_in_tool.exists():
            if copy_thumbnail_as_jpg(thumb_in_tool, dst_thumb):
                plog(f"Copied thumbnail: {dst_thumb.name}")
                thumb_copied = True
            else:
                plog(f"Cannot normalize thumbnail: {thumb_in_tool.name}", "WARN")

    # Fallback to global THUMB_DIR
    if not thumb_copied:
        thumb_path = find_thumbnail(code)
        if thumb_path:
            if copy_thumbnail_as_jpg(thumb_path, dst_thumb):
                plog(f"Copied thumbnail: {dst_thumb.name}")
                thumb_copied = True
            else:
                plog(f"Cannot normalize thumbnail: {thumb_path.name}", "WARN")

    # Root fallback: render succeeded, but no dedicated thumbnail exists.
    # DONE validation requires a JPG, so make one from project media instead of rerendering forever.
    if not thumb_copied:
        thumb_copied = create_fallback_thumbnail(project_info, dst_thumb, callback)
        if not thumb_copied:
            plog("No thumbnail source found; DONE validation will fail", "WARN")

    # 5. Copy thumb folder: VISUAL/<code>/thumb -> DONE/<code>/thumb
    visual_thumb_dir = project_dir / "thumb"
    if visual_thumb_dir.exists() and visual_thumb_dir.is_dir():
        dst_thumb_dir = done_folder / "thumb"
        shutil.copytree(visual_thumb_dir, dst_thumb_dir, dirs_exist_ok=True)
        plog(f"Copied thumb folder: {visual_thumb_dir} -> {dst_thumb_dir}")

    if not run_seo_tho_for_done(project_info, done_folder, callback):
        try:
            shutil.rmtree(done_folder)
        except Exception as e:
            plog(f"Cannot remove incomplete DONE folder: {e}", "WARN")
        return False, "SEO-tho failed"

    ok, missing = validate_done_folder(done_folder, require_thumb_folder=visual_thumb_dir.exists())
    if not ok:
        try:
            shutil.rmtree(done_folder)
        except Exception as e:
            plog(f"Cannot remove incomplete DONE folder: {e}", "WARN")
        return False, f"DONE folder missing: {', '.join(missing)}"

    files = list(done_folder.iterdir())
    plog(f"Done folder has {len(files)} files: {', '.join(f.name for f in files)}")

    return True, None


def delete_visual_project(project_info: Dict, callback=None) -> bool:
    code = project_info["code"]
    project_dir = project_info["path"]

    def plog(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            log(f"[{code}] {msg}", level)

    if not project_dir.exists():
        return True

    try:
        shutil.rmtree(project_dir)
        plog(f"Deleted VISUAL folder: {project_dir.name}")
        return True
    except Exception as e:
        plog(f"Cannot delete VISUAL folder: {e}", "WARN")
        return False


def cleanup_source_data(code: str, callback=None, max_retries: int = 3) -> bool:
    """Clean up source data after video is complete.

    Deletes (with retry):
    1. VISUAL/{code}/ folder
    2. Voice files from D:/AUTO/voice/{code}.*
    3. PROJECTS/{code}/ folder (last)
    """
    def plog(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            log(f"[{code}] {msg}", level)

    def safe_delete(path: Path, description: str) -> bool:
        """Delete file/folder with retry logic."""
        if not path.exists():
            return True

        for attempt in range(max_retries):
            try:
                if path.is_dir():
                    shutil.rmtree(path)
                else:
                    path.unlink()
                plog(f"Deleted {description}: {path.name}")
                return True
            except PermissionError:
                if attempt < max_retries - 1:
                    plog(f"  {description} in use, retry {attempt + 2}/{max_retries}...", "WARN")
                    time.sleep(2)  # Wait 2 seconds before retry
                else:
                    plog(f"Cannot delete {description}: file in use (will retry later)", "WARN")
            except Exception as e:
                plog(f"Cannot delete {description}: {e}", "WARN")
                break
        return False

    def is_voice_item_for_code(name: str, code_value: str) -> bool:
        """Match voice artifacts for a project code."""
        n = name.strip().lower()
        c = code_value.strip().lower()
        if not n or not c:
            return False
        return (
            n == c or
            n.startswith(c + ".") or
            n.startswith(c + "-") or
            n == f"srt_{c}" or
            n.startswith(f"srt_{c}.")
        )

    deleted_count = 0
    failed_items = []

    # 1. Delete from VISUAL folder
    visual_dir = VISUAL_DIR / code
    if visual_dir.exists():
        if safe_delete(visual_dir, "VISUAL folder"):
            deleted_count += 1
        else:
            failed_items.append(str(visual_dir))

    # 2. Delete from voice folder (files matching {code}.*)
    if VOICE_DIR.exists():
        # Delete files at root level (e.g. voice/KA1-0001.txt)
        for item in VOICE_DIR.iterdir():
            if item.is_file() and is_voice_item_for_code(item.name, code):
                if safe_delete(item, "voice file"):
                    deleted_count += 1
                else:
                    failed_items.append(str(item))

        # Delete inside template subfolders (e.g. voice/KA1-T3/KA1-0001.mp3)
        for subdir in VOICE_DIR.iterdir():
            if subdir.is_dir():
                for item in subdir.iterdir():
                    if is_voice_item_for_code(item.name, code):
                        if safe_delete(item, f"voice file ({subdir.name})"):
                            deleted_count += 1
                        else:
                            failed_items.append(str(item))

    # 3. Delete PROJECTS folder LAST (after VISUAL + voice)
    projects_dir = PROJECTS_DIR / code
    if projects_dir.exists():
        if safe_delete(projects_dir, "PROJECTS folder"):
            deleted_count += 1
        else:
            failed_items.append(str(projects_dir))

    if deleted_count > 0:
        plog(f"Cleanup complete: {deleted_count} items deleted")

    if failed_items:
        plog(f"Failed to delete {len(failed_items)} items (will retry next scan)", "WARN")

    return len(failed_items) == 0


def cleanup_leftover_done_projects() -> int:
    """Clean up leftover folders for projects that are already done.

    This catches cases where cleanup failed or was interrupted.
    Runs at the start of each scan cycle.
    """
    cleaned_count = 0

    if not DONE_DIR.exists():
        return 0

    with _processing_lock:
        active_codes = set(_processing_codes)

    # Get all codes that have complete DONE packages.
    # Do not treat a folder with only mp4 as done: copy_to_done creates DONE early,
    # then adds thumbnail/SEO later. Cleaning VISUAL during that window breaks edit.
    done_codes = set()
    for item in DONE_DIR.iterdir():
        if item.is_dir():
            if item.name in active_codes:
                continue
            ok, missing = validate_done_folder(item, require_thumb_folder=False)
            if ok:
                done_codes.add(item.name)
            elif list(item.glob("*.mp4")):
                log(f"  [CLEANUP] Skip incomplete DONE/{item.name}: missing {', '.join(missing)}", "WARN")

    if not done_codes:
        return 0

    # Check PROJECTS folder for leftover folders
    if PROJECTS_DIR.exists():
        for item in PROJECTS_DIR.iterdir():
            if item.is_dir() and item.name in done_codes:
                try:
                    shutil.rmtree(item)
                    log(f"  [CLEANUP] Removed leftover PROJECTS/{item.name}")
                    cleaned_count += 1
                except Exception as e:
                    log(f"  [CLEANUP] Cannot remove PROJECTS/{item.name}: {e}", "WARN")

    # Check VISUAL folder for leftover folders
    if VISUAL_DIR.exists():
        for item in VISUAL_DIR.iterdir():
            if item.is_dir() and item.name in done_codes:
                try:
                    shutil.rmtree(item)
                    log(f"  [CLEANUP] Removed leftover VISUAL/{item.name}")
                    cleaned_count += 1
                except Exception as e:
                    log(f"  [CLEANUP] Cannot remove VISUAL/{item.name}: {e}", "WARN")

    if cleaned_count > 0:
        log(f"  [CLEANUP] Total: {cleaned_count} leftover folders removed")

    return cleaned_count


# ============================================================================
# GOOGLE SHEET UPDATE
# ============================================================================

def load_gsheet_client():
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        log("gspread not installed", "ERROR")
        return None, None, None

    if not CONFIG_FILE.exists():
        return None, None, None

    try:
        cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        sa_path = cfg.get("SERVICE_ACCOUNT_JSON") or cfg.get("service_account_json")
        if not sa_path:
            return None, None, None

        spreadsheet_name = cfg.get("SPREADSHEET_NAME")
        if not spreadsheet_name:
            return None, None, None

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive.readonly",
        ]

        sa_file = Path(sa_path)
        if not sa_file.exists():
            sa_file = TOOL_DIR / "config" / sa_path

        if not sa_file.exists():
            return None, None, None

        creds = Credentials.from_service_account_file(str(sa_file), scopes=scopes)
        gc = gspread.authorize(creds)

        return gc, spreadsheet_name, cfg
    except Exception as e:
        log(f"Error loading gsheet client: {e}", "ERROR")
        return None, None, None


def update_sheet_status(codes: List[str], callback=None) -> Tuple[int, int]:
    if not codes:
        return 0, 0

    def plog(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            log(msg, level)

    gc, spreadsheet_name, cfg = load_gsheet_client()
    if not gc:
        plog("Sheet update skipped: không load được Google Sheet client (thiếu creds.json hoặc gspread?)", "WARN")
        return -1, 0

    try:
        from gspread.exceptions import APIError

        def do_update():
            ws = gc.open(spreadsheet_name).worksheet(SOURCE_SHEET_NAME)
            raw_g = ws.col_values(SOURCE_COL_CODE)
            raw_m = ws.col_values(SOURCE_COL_STATUS)

            code_to_rows = {}
            for idx, val in enumerate(raw_g, start=1):
                norm = normalize_code(val)
                if norm:
                    code_to_rows.setdefault(norm, []).append(idx)

            targets = [normalize_code(c) for c in codes if c]
            targets = list(set(t for t in targets if t))

            plog(f"Updating {len(targets)} codes in sheet...")

            found, updates = 0, []
            for code in targets:
                rows = code_to_rows.get(code, [])
                if not rows:
                    continue

                found += len(rows)
                for r in rows:
                    current = raw_m[r-1] if r-1 < len(raw_m) else ""
                    if current.strip().upper() == STATUS_VALUE.upper():
                        continue

                    plog(f"  Updating {code} @ row {r}")
                    updates.append({"range": f"M{r}", "values": [[STATUS_VALUE]]})

            if not updates:
                return found, 0

            ws.batch_update(updates, value_input_option="USER_ENTERED")
            plog(f"Updated {len(updates)} rows")
            return found, len(updates)

        last_error = None
        for attempt in range(MAX_RETRIES):
            try:
                return do_update()
            except APIError as e:
                last_error = e
                if e.response.status_code in (429, 500, 502, 503, 504):
                    delay = RETRY_BASE_DELAY * (2 ** attempt)
                    plog(f"API error {e.response.status_code}, retrying in {delay}s...", "WARN")
                    time.sleep(delay)
                else:
                    raise
            except Exception as e:
                last_error = e
                if "timeout" in str(e).lower() or "connection" in str(e).lower():
                    delay = RETRY_BASE_DELAY * (2 ** attempt)
                    plog(f"Network error, retrying in {delay}s...", "WARN")
                    time.sleep(delay)
                else:
                    raise

        raise last_error

    except Exception as e:
        plog(f"Error updating sheet: {e}", "ERROR")
        return 0, 0


# ============================================================================
# PROCESS PROJECT
# ============================================================================

def generate_thumbnail_for_project(project_info: Dict, callback=None) -> bool:
    """Generate thumbnail for project if thumb folder exists."""
    code = project_info["code"]
    project_dir = project_info["path"]

    def plog(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            log(f"[{code}] {msg}", level)

    # Check for thumbnail folder in VISUAL project. Some producers use "thumb",
    # older ones use "thumbnail".
    thumbnail_folder = project_dir / "thumbnail"
    if not thumbnail_folder.exists():
        thumbnail_folder = project_dir / "thumb"
    if not thumbnail_folder.exists():
        return False

    # Find source image (prefer thumb_003.png, fallback to any image)
    valid_ext = {".png", ".jpg", ".jpeg", ".webp"}
    source_img = thumbnail_folder / "thumb_003.png"
    if not source_img.exists():
        source_img = None
        for f in thumbnail_folder.iterdir():
            if f.is_file() and f.suffix.lower() in valid_ext:
                source_img = f
                break

    if not source_img:
        plog("  No source image in thumbnail folder", "WARN")
        return False

    plog(f"  Found thumb source: {source_img.name}")

    # Run thumbnail generation
    try:
        thumb_script = TOOL_DIR / "run_thumb.py"
        if not thumb_script.exists():
            plog("  run_thumb.py not found", "WARN")
            return False

        import subprocess
        result = subprocess.run(
            [sys.executable, str(thumb_script), code],
            capture_output=True, text=True, timeout=300,
            creationflags=SUBPROCESS_FLAGS if sys.platform == "win32" else 0
        )

        if result.returncode == 0:
            plog("  Thumbnail generated successfully")
            return True
        else:
            plog(f"  Thumbnail generation failed: {result.stderr[-200:] if result.stderr else 'Unknown error'}", "WARN")
            return False

    except subprocess.TimeoutExpired:
        plog("  Thumbnail generation timed out", "WARN")
        return False
    except Exception as e:
        plog(f"  Thumbnail error: {e}", "WARN")
        return False


def process_project(project_info: Dict, callback=None) -> bool:
    code = project_info["code"]
    project_dir = project_info["path"]

    def plog(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            log(f"[{code}] {msg}", level)

    # Mark as processing (prevent duplicate processing)
    with _processing_lock:
        if code in _processing_codes:
            plog("Already being processed by another worker, skipping...", "WARN")
            return False
        _processing_codes.add(code)

    try:
        plog("="*50)
        plog(f"Processing: {code}")
        plog("="*50)

        # Check if folder is stable (VM finished copying)
        update_progress(code=code, step="Checking folder", percent=0, status="composing")
        plog("Checking if folder is stable (VM copy complete)...")
        if not is_folder_stable(project_dir):
            plog("Folder is still being copied, skipping for now...", "WARN")
            return False

        plog(f"Media: {project_info['video_count']} videos + {project_info['image_count']} images")
        plog(f"Scenes: {project_info['total_scenes']}")

        success, video_path, error = compose_video(project_info, callback)
        if not success:
            plog(f"Compose failed: {error}", "ERROR")
            return False

        # Generate legacy thumbnail output before DONE copy so SEO-tho can rename
        # the final jpg once, after all assets are present.
        generate_thumbnail_for_project(project_info, callback)

        success, error = copy_to_done(project_info, video_path, callback)
        if not success:
            plog(f"Copy failed: {error}", "ERROR")
            return False

        # Update sheet status with aggressive retry (MUST succeed before cleanup)
        sheet_updated = False
        max_sheet_retries = 10
        for sheet_attempt in range(max_sheet_retries):
            found, updated = update_sheet_status([code], callback)
            if found == -1:
                # Config error (missing creds.json / gspread) - no point retrying
                plog(f"Sheet update skipped: thiếu config. Copy config/creds.json từ máy chính.", "ERROR")
                break
            if updated > 0:
                plog(f"Sheet updated: {STATUS_VALUE}")
                sheet_updated = True
                break
            elif found > 0:
                # Found but already had correct status
                plog(f"Sheet already has correct status")
                sheet_updated = True
                break
            else:
                # Not found or error - retry
                if sheet_attempt < max_sheet_retries - 1:
                    delay = min(30, 5 * (sheet_attempt + 1))  # 5s, 10s, 15s... max 30s
                    plog(f"Sheet update failed, retrying in {delay}s... (attempt {sheet_attempt + 1}/{max_sheet_retries})", "WARN")
                    time.sleep(delay)

        if not sheet_updated:
            plog(f"Sheet update failed for {code}. Video done nhưng chưa điền Sheet.", "ERROR")
            # Still return True because video is done, but don't cleanup
            return True

        # Clean up source data (voice folder + PROJECTS folder) ONLY after sheet is updated
        cleanup_source_data(code, callback)

        plog(f"DONE: {code}")
        return True

    finally:
        # Always remove from processing set (whether success or failure)
        with _processing_lock:
            _processing_codes.discard(code)
        # Remove from progress tracking
        update_progress(code=code, remove=True)


# ============================================================================
# MAIN
# ============================================================================

def run_scan_loop(parallel: int = None):
    # Detect system resources
    resources = get_system_resources()
    global CLIP_WORKERS
    CLIP_WORKERS = get_optimal_workers(resources, "clip")

    # Auto-detect optimal parallel if not specified
    if parallel is None:
        parallel = get_optimal_workers(resources, "parallel")
        parallel_mode = "auto"
    else:
        parallel_mode = "manual"

    log("="*60)
    log("  VE3 TOOL - EDIT MODE (Compose MP4)")
    log("="*60)
    log(f"  VISUAL folder: {VISUAL_DIR}")
    log(f"  DONE folder:   {DONE_DIR}")
    log(f"  Scan interval: {SCAN_INTERVAL}s")
    log("-"*60)
    log("  HARDWARE DETECTED:")
    log(f"    CPU cores:   {resources['cpu_physical']} physical / {resources['cpu_cores']} logical")
    log(f"    RAM:         {resources['ram_gb']:.1f} GB")
    log(f"    GPU:         {'NVENC Available' if resources['gpu_available'] else 'Not available (using CPU)'}")
    log("-"*60)
    log("  AUTO-OPTIMIZED SETTINGS:")
    log(f"    Parallel videos: {parallel} ({parallel_mode})")
    log(f"    Clip workers:    {CLIP_WORKERS} per video")
    log(f"    Total threads:   ~{parallel * CLIP_WORKERS} max")
    log("="*60)

    # Set hardware info in progress for GUI
    set_hardware_info({
        "cpu_cores": resources['cpu_cores'],
        "cpu_physical": resources['cpu_physical'],
        "ram_gb": round(resources['ram_gb'], 1),
        "gpu": "NVENC" if resources['gpu_available'] else "CPU",
        "clip_workers": CLIP_WORKERS,
        "parallel": parallel,
    })

    # Clear any stale progress on start
    with _progress_lock:
        _multi_progress["videos"] = {}
        _write_progress()

    DONE_DIR.mkdir(parents=True, exist_ok=True)

    MAX_FAILURES = 3  # Skip after this many consecutive failures
    fail_counts = {}        # code -> consecutive failure count
    submitted_codes = set() # codes currently queued or running in executor
    active_futures = {}     # future -> project_info
    cycle = 0

    # Slot queue: slot 0 = CPU subtitle, slot 1 = NVENC subtitle
    # Workers grab a slot on start, release on finish → always N parallel
    _slot_q = _queue_module.Queue()
    for _i in range(parallel):
        _slot_q.put(_i)

    def _process_with_slot(project_info):
        """Acquire a slot, process, release slot. Called by thread pool."""
        slot = _slot_q.get()
        p = dict(project_info)
        p['_slot'] = slot
        p['_parallel'] = parallel
        try:
            return process_project(p)
        finally:
            _slot_q.put(slot)

    with ThreadPoolExecutor(max_workers=parallel) as executor:
        while True:
            cycle += 1
            log(f"\n[CYCLE {cycle}] Scanning VISUAL folder...")

            cleanup_leftover_done_projects()
            all_pending = scan_visual_projects()

            # Filter failures and already-queued
            skipped = []
            new_to_submit = []
            for p in all_pending:
                code = p["code"]
                if fail_counts.get(code, 0) >= MAX_FAILURES:
                    skipped.append(code)
                elif code not in submitted_codes:
                    new_to_submit.append(p)

            if skipped:
                log(f"  [SKIP] {len(skipped)} broken projects (failed {MAX_FAILURES}+ times): {', '.join(skipped)}", "WARN")

            total_visible = len(new_to_submit) + len(submitted_codes)
            if new_to_submit:
                log(f"  Found {total_visible} project(s) ready ({len(submitted_codes)} already running, {len(new_to_submit)} new):")
                for p in new_to_submit[:5]:
                    priority = project_priority_key(p)[0]
                    log(f"    - {p['code']} ({p['video_count']}v + {p['image_count']}i / {p['total_scenes']} scenes, priority={priority:.1f})")
                if len(new_to_submit) > 5:
                    log(f"    ... and {len(new_to_submit) - 5} more")
                # Submit new projects to rolling pool (executor queues them, runs up to parallel at once)
                for p in new_to_submit:
                    submitted_codes.add(p['code'])
                    f = executor.submit(_process_with_slot, p)
                    active_futures[f] = p
            elif not active_futures:
                log("  No pending projects")
                update_progress(code="", step="Waiting", percent=0, clip_current=0, clip_total=0, status="idle")

            if active_futures:
                # Wait for at least one completion (or SCAN_INTERVAL timeout to rescan)
                done, _ = _cf.wait(
                    list(active_futures.keys()),
                    return_when=_cf.FIRST_COMPLETED,
                    timeout=SCAN_INTERVAL
                )
                for f in done:
                    project = active_futures.pop(f)
                    code = project["code"]
                    submitted_codes.discard(code)
                    try:
                        success = f.result()
                        if success:
                            log(f"  {code}: SUCCESS", "OK")
                            fail_counts.pop(code, None)
                        else:
                            fail_counts[code] = fail_counts.get(code, 0) + 1
                            cnt = fail_counts[code]
                            if cnt < MAX_FAILURES:
                                log(f"  {code}: FAILED (retry {cnt}/{MAX_FAILURES})", "ERROR")
                            else:
                                log(f"  {code}: FAILED {MAX_FAILURES} times - SKIPPING", "ERROR")
                    except Exception as e:
                        fail_counts[code] = fail_counts.get(code, 0) + 1
                        log(f"  {code}: EXCEPTION - {e}", "ERROR")
                update_progress()
                # Loop immediately to scan for newly added projects
                continue

            # Nothing active, nothing new - wait before next scan
            log(f"\n  Waiting {SCAN_INTERVAL}s... (Ctrl+C to stop)")
            try:
                time.sleep(SCAN_INTERVAL)
            except KeyboardInterrupt:
                log("\n\nStopped by user.")
                break


def run_single_project(code: str):
    project_dir = VISUAL_DIR / code

    if not project_dir.exists():
        log(f"Project not found: {project_dir}", "ERROR")
        return

    info = get_project_info(project_dir)

    if info["already_done"]:
        log(f"Project already done: {code}", "WARN")
        return

    if not info["ready_for_edit"]:
        log(f"Project not ready: {code}", "WARN")
        log(f"  Media: {info['video_count']}v + {info['image_count']}i / {info['total_scenes']} scenes")
        log(f"  Audio: {info['has_audio']}")
        log(f"  Excel: {info['has_excel']}")
        return

    process_project(info)


def run_scan_only():
    log("="*60)
    log("  VE3 TOOL - EDIT MODE (Scan Only)")
    log("="*60)

    pending = scan_visual_projects()

    if not pending:
        log("No pending projects found")
        return

    log(f"\nFound {len(pending)} projects ready to edit:\n")

    for p in pending:
        log(f"  {p['code']}:")
        log(f"    Media:  {p['video_count']} videos + {p['image_count']} images / {p['total_scenes']} scenes")
        log(f"    Audio:  {'YES' if p['has_audio'] else 'NO'}")
        log(f"    Excel:  {'YES' if p['has_excel'] else 'NO'}")
        log(f"    SRT:    {'YES' if p['has_srt'] else 'NO'}")


def main():
    parser = argparse.ArgumentParser(description="VE3 Tool - Edit Mode (Compose MP4)")
    parser.add_argument("code", nargs="?", help="Process single project by code")
    parser.add_argument("--parallel", "-p", type=int, default=None,
                        help="Number of parallel videos (default: auto-detect based on hardware)")
    parser.add_argument("--scan-only", action="store_true", help="Only scan and show status")
    args = parser.parse_args()

    if args.scan_only:
        run_scan_only()
    elif args.code:
        run_single_project(args.code)
    else:
        run_scan_loop(parallel=args.parallel)


if __name__ == "__main__":
    main()
